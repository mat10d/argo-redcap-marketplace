#!/usr/bin/env python3
"""End-to-end test of the QA worklist builder against the synthetic study.

This test exists because of a bug that every other kind of test missed: after the 0.9.0
refactor, missing_and_certainty_for returned bare bools on four paths into a tuple-unpacking
caller — so the builder crashed the moment any applicable-but-blank cell existed, meaning the
shipped plugin could not produce a nonempty worklist AT ALL. Unit tests of the branching logic
all passed; nothing ran the builder on data with flaggable cells until the synthetic fixture
did, on first contact.

Asserts against MANIFEST.json's engineered counts — numbers, not vibes.
"""
from __future__ import annotations

import importlib.util
import json
import re
import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
FIXTURE = REPO / "testing" / "fixtures" / "synthetic-study"
QA_SKILL = REPO / "plugins" / "argo-qa-specialist" / "skills" / "qa-worklists"
BUILDER = QA_SKILL / "build_worklists.py"
SKILL_DOC = QA_SKILL / "SKILL.md"

# The highlight colours come from the skill's own single definition, never retyped in a test.
sys.path.insert(0, str(QA_SKILL))

try:
    import openpyxl  # noqa: F401
    import pandas  # noqa: F401
    import yaml  # noqa: F401
    DEPS = True
except ImportError:
    DEPS = False


@unittest.skipIf(not DEPS, "pandas/openpyxl/yaml not installed")
class TestWorklistBuilderEndToEnd(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.out = Path(tempfile.mkdtemp())
        cls.proc = subprocess.run(
            [sys.executable, str(BUILDER),
             "--records-csv", str(FIXTURE / "records.csv"),
             "--metadata-csv", str(FIXTURE / "datadictionary.csv"),
             "--fields", str(FIXTURE / "qa_fields.yaml"),
             "--out", str(cls.out), "--id-field", "syn_id"],
            capture_output=True, text=True, timeout=300,
        )
        cls.manifest = json.loads((FIXTURE / "MANIFEST.json").read_text())

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.out, ignore_errors=True)

    def test_builder_completes_without_crashing(self):
        self.assertEqual(self.proc.returncode, 0,
                         f"builder crashed on the fixture:\n{self.proc.stderr[-1500:]}")
        self.assertNotIn("Traceback", self.proc.stdout + self.proc.stderr)

    def test_all_eight_workbooks_written(self):
        # 2 workbooks x 2 DAGs x (with_MDC + no_MDC)
        books = list(self.out.rglob("*.xlsx"))
        self.assertEqual(len(books), 8, f"expected 8 workbooks, got {[b.name for b in books]}")

    def test_unparseable_datediff_reported_exactly_once(self):
        combined = self.proc.stdout + self.proc.stderr
        self.assertEqual(combined.count("datediff"), 1,
                         "the one engineered unreadable condition must be reported once")

    def test_row_counts_match_the_manifest(self):
        """The fixture's MANIFEST states exact expected rows per workbook — assert them."""
        expected = self.manifest.get("qa", {}).get("expected_worklists", {})
        self.assertTrue(expected, "MANIFEST lost its qa.expected_worklists block")
        for workbook, variants in expected.items():
            for key, spec in variants.items():
                site, variant = key.split("/")
                matches = [p for p in self.out.rglob(f"{workbook}_{site}.xlsx")
                           if p.parent.name == variant]
                self.assertTrue(matches, f"{variant}/{workbook}_{site}.xlsx not produced")
                wb = openpyxl.load_workbook(matches[0])
                ws = wb.active
                data_rows = ws.max_row - 2  # header + prereq rows
                self.assertEqual(
                    data_rows, spec["rows_with_work"],
                    f"{workbook} {key}: {data_rows} data rows, "
                    f"MANIFEST says {spec['rows_with_work']}")


@unittest.skipIf(not DEPS, "pandas/openpyxl/yaml not installed")
class TestDuplicateFieldLabels(unittest.TestCase):
    """0.19 #41: the builder crashed outright on any dictionary that reuses a field label.

    REDCap requires field NAMES to be unique; labels are free to repeat, and in practice they
    do — the live colorectal dictionary has 44 labels shared across 160 fields ("Date", "Other,
    specify", "Result"). The builder renamed its dataframe's columns to labels and then read
    each cell back by label, so a shared label produced two identically-named columns, pandas
    handed back a Series where a value was expected, and openpyxl killed the run with
    `ValueError: Cannot convert ... to Excel`. Nobody saw it for a while only because the
    colliding fields on the first project that hit it were @HIDDEN.

    Two things are asserted, because fixing only the first would trade a crash for an
    unreadable workbook: the build completes, AND the two fields arrive in two columns under
    two different headings, each carrying its own field's values.
    """

    DD = (
        "field_name,form_name,section_header,field_type,field_label,"
        "select_choices_or_calculations,field_note,"
        "text_validation_type_or_show_slider_number,text_validation_min,text_validation_max,"
        "identifier,branching_logic,required_field,custom_alignment,question_number,"
        "matrix_group_name,matrix_ranking,field_annotation\n"
        "record_id,f1,,text,Record ID,,,,,,,,,,,,,\n"
        "biopsy,f1,,radio,Biopsy taken,\"1, Yes | 0, No\",,,,,,,,,,,,\n"
        # The collision: two different fields, one label, both gated by the same question.
        "first_date,f1,,text,Date of procedure,,,date_ymd,,,,[biopsy] = '1',,,,,,\n"
        "second_date,f1,,text,Date of procedure,,,date_ymd,,,,[biopsy] = '1',,,,,,\n"
    )
    RECORDS = (
        "record_id,redcap_data_access_group,biopsy,first_date,second_date\n"
        "1,alpha,1,,\n"
        "2,alpha,1,2024-03-01,\n"
        "3,alpha,1,,2024-05-09\n"
    )
    CFG = ("workbooks:\n  - name: dup\n    title: Dup\n"
           "    fields: [first_date, second_date]\n")

    @classmethod
    def setUpClass(cls):
        cls.tmp = Path(tempfile.mkdtemp())
        (cls.tmp / "dd.csv").write_text(cls.DD)
        (cls.tmp / "records.csv").write_text(cls.RECORDS)
        (cls.tmp / "fields.yaml").write_text(cls.CFG)
        cls.proc = subprocess.run(
            [sys.executable, str(BUILDER),
             "--records-csv", str(cls.tmp / "records.csv"),
             "--metadata-csv", str(cls.tmp / "dd.csv"),
             "--fields", str(cls.tmp / "fields.yaml"),
             "--out", str(cls.tmp / "out"), "--round="],
            capture_output=True, text=True, timeout=300)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmp, ignore_errors=True)

    def sheet(self):
        return openpyxl.load_workbook(
            self.tmp / "out" / "with_MDC" / "dup_alpha.xlsx").active

    def test_the_build_completes(self):
        self.assertEqual(self.proc.returncode, 0,
                         f"a shared field label must not stop the build:\n"
                         f"{self.proc.stdout[-800:]}\n{self.proc.stderr[-1500:]}")
        self.assertNotIn("Cannot convert", self.proc.stderr)
        self.assertNotIn("Traceback", self.proc.stdout + self.proc.stderr)

    def test_both_fields_get_their_own_column_under_distinct_headings(self):
        headers = [c.value for c in self.sheet()[1]]
        self.assertEqual(len(headers), len(set(headers)),
                         f"two columns share a heading, so the RA cannot tell them apart: "
                         f"{headers}")
        # The gate column comes first, then the two colliding fields, then RESPONSE.
        self.assertEqual(headers,
                         ["record_id", "Biopsy taken", "Date of procedure",
                          "Date of procedure (second_date)", "RESPONSE"])

    def test_each_column_carries_its_own_fields_values(self):
        """The crash is the loud half; reading the WRONG field's value would be the quiet one."""
        ws = self.sheet()
        headers = [c.value for c in ws[1]]
        first = headers.index("Date of procedure") + 1
        second = headers.index("Date of procedure (second_date)") + 1
        rows = {str(ws.cell(row=r, column=1).value): (
                    str(ws.cell(row=r, column=first).value or ""),
                    str(ws.cell(row=r, column=second).value or ""))
                for r in range(3, ws.max_row + 1)}
        self.assertEqual(rows["2"], ("2024-03-01", ""), "first_date's value under first_date")
        self.assertEqual(rows["3"], ("", "2024-05-09"), "second_date's value under second_date")

    def test_the_right_cells_are_flagged_in_each_column(self):
        """Both fields are gated by [biopsy]='1' and blank where they're blank."""
        from qa_colours import YELLOW_HEX
        ws = self.sheet()
        headers = [c.value for c in ws[1]]
        painted = {(str(ws.cell(row=r, column=1).value), headers[c - 1])
                   for r in range(3, ws.max_row + 1)
                   for c in range(2, ws.max_column + 1)
                   if str(ws.cell(row=r, column=c).fill.fgColor.rgb or "").upper()
                   .endswith(YELLOW_HEX)}
        self.assertEqual(painted, {
            ("1", "Date of procedure"), ("1", "Date of procedure (second_date)"),
            ("2", "Date of procedure (second_date)"), ("3", "Date of procedure"),
        })


class TestDisplayHeaders(unittest.TestCase):
    """Unit-level: headings are display only, and always distinct."""

    @classmethod
    def setUpClass(cls):
        spec = importlib.util.spec_from_file_location("_argo_build_worklists", BUILDER)
        cls.bw = importlib.util.module_from_spec(spec)
        sys.modules["build_worklists"] = cls.bw
        try:
            spec.loader.exec_module(cls.bw)
        except ImportError as e:
            raise unittest.SkipTest(f"build_worklists dependencies unavailable: {e}")

    def test_unique_labels_are_left_alone(self):
        self.assertEqual(
            self.bw.display_headers(["a", "b"], {"a": "Age", "b": "Sex"}), ["Age", "Sex"])

    def test_a_repeat_gets_its_field_name(self):
        self.assertEqual(
            self.bw.display_headers(["a", "b", "c"], {"a": "Date", "b": "Date", "c": "Date"}),
            ["Date", "Date (b)", "Date (c)"])

    def test_a_field_with_no_label_falls_back_to_its_name(self):
        self.assertEqual(self.bw.display_headers(["odd_field"], {}), ["odd_field"])

    def test_id_columns_are_not_shadowed(self):
        """An ID column sits to the left; a data column may not repeat its heading."""
        got = self.bw.display_headers(["x"], {"x": "record_id"}, taken=["record_id"])
        self.assertEqual(got, ["record_id (x)"])

    def test_headings_are_always_distinct(self):
        """Even the pathological case: a plain label that already looks disambiguated."""
        got = self.bw.display_headers(
            ["a", "b"], {"a": "Date (b)", "b": "Date"})
        self.assertEqual(len(got), len(set(got)), got)


class TestScopeFirstIsDocumented(unittest.TestCase):
    """0.19 #45: the plan step must START by asking what to QA.

    Live dictionaries run to hundreds of fields. A session that proposes "every chaseable
    blank" hands the RAs a wall of yellow and buries the ten fields the QA specialist actually
    wanted, so the scope question comes first and a vague answer gets drilled down rather than
    interpreted.
    """

    DOC = SKILL_DOC.read_text()

    def plan_section(self):
        """The plan section, whitespace collapsed and lowercased — the doc wraps its lines."""
        self.assertIn("### The workbook plan", self.DOC)
        section = self.DOC.split("### The workbook plan", 1)[1].split("\n### ", 1)[0]
        return re.sub(r"\s+", " ", section).lower()

    def test_the_section_asks_what_to_qa_before_proposing_anything(self):
        plan = self.plan_section()
        self.assertIn("what exactly do you want me to qa", plan)
        ask = plan.index("what exactly do you want me to qa")
        propose = plan.index("workbooks, the fields in each")
        self.assertLess(ask, propose,
                        "the scope question must come before the workbook proposal")

    def test_a_vague_answer_is_narrowed_against_the_dictionary(self):
        plan = self.plan_section()
        self.assertIn("staging", plan,
                      "the doc should work the vague-answer case through an example")
        self.assertRegex(plan, r"data dictionary.{0,120}list the fields")
        self.assertIn("narrow", plan)

    def test_chasing_everything_is_ruled_out_as_a_default(self):
        plan = self.plan_section()
        self.assertIn("never the default plan", plan)
        self.assertRegex(plan, r"600\+? fields",
                         "say plainly how big real dictionaries get")

    def test_still_exactly_one_confirming_question_on_the_split(self):
        """#16 stays true: the user never writes the yaml, and confirms with one question."""
        plan = self.plan_section()
        self.assertIn("**one** question", plan)
        self.assertIn("never asked to produce it", plan)
        self.assertIn("not homework", plan)


if __name__ == "__main__":
    unittest.main(verbosity=2)
