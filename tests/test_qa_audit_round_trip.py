#!/usr/bin/env python3
"""End-to-end test of the QA specialist's TASK 2 — auditing RA-returned worklists.

Task 1 (build the worklists) has had a test since the synthetic fixture landed. Task 2 had
none at all: `review_responses.py` had never once been run against a workbook with RA edits in
it by anything automated. This closes that hole.

The round trip:

    records.csv + datadictionary.csv + qa_fields.yaml
        -> build_worklists.py                     (the worklists we'd send the sites)
        -> generate_returns.py                    (seeded, engineered RA edits)
        -> review_responses.py                    (the audit)
        -> summarize_for_ra.py                    (what goes back to each RA)
        -> counts asserted against MANIFEST.json's `returned` block

Nothing binary is committed: the workbooks are generated into a temp directory on every run
from the seeded generator, and thrown away afterwards.

Four of the assertions here used to PIN KNOWN DEFECTS — three in `review_responses.py`, one in
`build_worklists.py` — and asserted the broken behaviour so the gaps stayed visible in the
suite instead of being discovered by an RA whose answers were silently dropped. All four were
fixed in 0.17.2 and the assertions were flipped to the correct behaviour:

    amber answers are reported (and tagged amber)      test_amber_cell_answers_*
    out-of-scope edits are reported separately         test_out_of_scope_edits_*
    a single-ID workbook loses nothing                 test_single_id_column_workbook_*
    gate-context column order is stable                test_gate_context_column_order_*
"""
from __future__ import annotations

import csv
import importlib.util
import json
import os
import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
FIXTURE = REPO / "testing" / "fixtures" / "synthetic-study"
GENERATOR = FIXTURE / "generate_returns.py"
QA_SKILL = REPO / "plugins" / "argo-qa-specialist" / "skills" / "qa-worklists"
REVIEWER = QA_SKILL / "review_responses.py"
BUILDER = QA_SKILL / "build_worklists.py"
SUMMARIZER = QA_SKILL / "summarize_for_ra.py"

# The highlight colours come from the skill's own single definition, never retyped here: a test
# that hunts for a hex the builder stopped painting finds nothing and asserts nothing.
sys.path.insert(0, str(QA_SKILL))
from qa_colours import AMBER_HEX, LEGACY_FLAG_HEXES, YELLOW_HEX  # noqa: E402

try:
    import openpyxl  # noqa: F401
    import pandas  # noqa: F401
    import yaml  # noqa: F401
    DEPS = True
except ImportError:
    DEPS = False


def _load_reviewer():
    """Import review_responses.py by path — the plugin folder never goes on sys.path."""
    spec = importlib.util.spec_from_file_location("_argo_review_responses", REVIEWER)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


@unittest.skipIf(not DEPS, "pandas/openpyxl/yaml not installed")
class TestQAAuditRoundTrip(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.out = Path(tempfile.mkdtemp())
        cls.proc = subprocess.run(
            [sys.executable, str(GENERATOR), "--out", str(cls.out)],
            capture_output=True, text=True, timeout=600,
        )
        cls.manifest = json.loads((FIXTURE / "MANIFEST.json").read_text())
        cls.block = cls.manifest.get("returned", {})
        cls.rr = _load_reviewer()

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.out, ignore_errors=True)

    # -- helpers ----------------------------------------------------------

    def pair(self, name):
        return (self.out / "build" / "with_MDC" / f"{name}.xlsx",
                self.out / "returned" / f"{name}_RETURNED.xlsx")

    def expectations(self):
        exp = self.block.get("expected_review_responses")
        self.assertTrue(exp, "MANIFEST.json has no `returned.expected_review_responses` block — "
                             "run: python3 testing/fixtures/synthetic-study/generate_returns.py "
                             "--out /tmp/x --update-manifest")
        return exp

    def audit(self, name):
        orig, resp = self.pair(name)
        return self.rr.diff(str(orig), str(resp))

    # -- the generator ----------------------------------------------------

    def test_generator_completes_without_crashing(self):
        self.assertEqual(self.proc.returncode, 0,
                         f"generate_returns.py failed:\n{self.proc.stdout[-1500:]}"
                         f"\n{self.proc.stderr[-1500:]}")
        self.assertNotIn("Traceback", self.proc.stdout + self.proc.stderr)

    def test_four_returned_workbooks_written(self):
        books = sorted(p.name for p in (self.out / "returned").glob("*.xlsx"))
        self.assertEqual(len(books), 4, f"expected 4 returned workbooks, got {books}")

    def test_manifest_block_matches_what_the_generator_just_produced(self):
        """MANIFEST's stated counts must be the counts the seeded generator actually makes.

        If this fails, the fixture or the builder changed: rerun generate_returns.py with
        --update-manifest and check the diff is what you intended.
        """
        fresh = json.loads((self.out / "returned" / "returned_counts.json").read_text())
        self.assertTrue(self.block, "MANIFEST.json has no `returned` block")
        self.assertEqual(self.block.get("per_file"), fresh["per_file"])
        self.assertEqual(self.block.get("totals"), fresh["totals"])

    # -- the audit script -------------------------------------------------

    def test_reviewer_runs_on_every_returned_workbook(self):
        for name in self.expectations():
            orig, resp = self.pair(name)
            with self.subTest(name):
                proc = subprocess.run([sys.executable, str(REVIEWER), str(orig), str(resp)],
                                      capture_output=True, text=True, timeout=300)
                self.assertEqual(proc.returncode, 0,
                                 f"review_responses.py failed on {name}:\n{proc.stderr[-1500:]}")
                self.assertNotIn("Traceback", proc.stdout + proc.stderr)
                self.assertIn("RESPONSE column present: True", proc.stdout)

    def test_triage_counts_match_the_manifest(self):
        """The four-bucket triage inputs, per returned workbook, as exact numbers.

        READY / QUESTION_FOR_RA are both sourced from the answered-cell list (the specialist
        splits them by whether the value maps to the DD); NO_ACTION / VERIFY are both sourced
        from the note-only list. Out-of-scope edits are a fifth thing entirely — nobody asked
        for them — and are asserted separately below.
        """
        for name, spec in self.expectations().items():
            with self.subTest(name):
                a = self.audit(name)
                yellow = sum(1 for cells in a.by_record.values()
                             for ans in cells if ans.kind == "yellow")
                amber = sum(1 for cells in a.by_record.values()
                            for ans in cells if ans.kind == "amber")
                note_only = [rid for rid, n in a.notes.items() if n and rid not in a.by_record]
                self.assertEqual(a.has_response_col, spec["response_column_present"])
                self.assertEqual(a.id_field, "syn_id")
                self.assertEqual(len(a.by_record), spec["records_with_proposed_updates"],
                                 f"{name}: records with proposed updates")
                self.assertEqual(yellow, spec["changed_cells"], f"{name}: yellow cells answered")
                self.assertEqual(amber, spec["amber_cells_detected"],
                                 f"{name}: amber cells answered")
                self.assertEqual(len(note_only), spec["note_only_records"],
                                 f"{name}: records with a note but no cell change")

    def test_note_only_records_are_the_engineered_ones(self):
        """The VERIFY/NO_ACTION bucket must name the right patients, not just the right count."""
        for name, counts in self.block["per_file"].items():
            with self.subTest(name):
                a = self.audit(name)
                note_only = sorted(rid for rid, n in a.notes.items()
                                   if n and rid not in a.by_record)
                self.assertEqual(note_only, counts["note_only_record_ids"])

    def test_mdc_answers_survive_the_audit_as_changed_cells(self):
        """An RA answering with an MDC sentinel is an answer, not a blank — it must show up."""
        for name, counts in self.block["per_file"].items():
            if not counts["filled_with_mdc"]:
                continue
            with self.subTest(name):
                a = self.audit(name)
                seen = [ans.now for cells in a.by_record.values() for ans in cells]
                mdc_seen = sum(1 for v in seen if v in self.manifest["qa"]["sentinel_codes"])
                self.assertEqual(mdc_seen, counts["filled_with_mdc"],
                                 f"{name}: MDC answers reported")

    # -- the four formerly-pinned defects ---------------------------------

    def test_out_of_scope_edits_are_reported_separately(self):
        """An RA edit to a cell nobody flagged must be reported — as its own thing.

        Was a silent hole: `_yellow_keys` only recorded yellow cells and `diff` only looked up
        response values for those keys, so a value the RA overwrote in a gate-context column
        (e.g. changing 'Sex' from Female to Male) passed the audit completely silently, while
        SKILL.md claimed the audit shows every changed cell. The fixture engineers 9 such edits
        across the four workbooks.

        They must NOT be folded into the answers: an unrequested edit is not an answer to a
        question we asked, and can change which other fields even apply.
        """
        total_engineered = total_detected = 0
        for name, counts in self.block["per_file"].items():
            a = self.audit(name)
            detected = {(e.record, e.field) for e in a.out_of_scope}
            for edit in counts["out_of_scope_detail"]:
                total_engineered += 1
                key = (edit["record"], edit["column"])
                with self.subTest(name=name, edit=key):
                    self.assertIn(key, detected, f"{name}: out-of-scope edit not reported")
                    answered = [ans.field for ans in a.by_record.get(edit["record"], [])]
                    self.assertNotIn(edit["column"], answered,
                                     "an unrequested edit must not be reported as an answer")
                total_detected += 1
            # and nothing invented: exactly the engineered edits, no more
            with self.subTest(name):
                self.assertEqual(len(a.out_of_scope), counts["out_of_scope_edits"],
                                 f"{name}: out-of-scope edits reported")
        self.assertEqual(total_engineered, 9, "fixture should engineer 9 out-of-scope edits")
        self.assertEqual(total_detected, 9, "all 9 out-of-scope edits must be reported")

    def test_amber_cell_answers_are_reported_and_tagged(self):
        """An answer in an AMBER cell is an answer, and is labelled as amber.

        Amber means 'we could not read this field's condition — please check'. It is a genuine
        RA task, documented as such in SKILL.md, but the audit used to match only the yellow
        fill so every amber answer was dropped. The fixture fills 5 amber cells.

        Tagging matters as much as counting: an amber answer needs the extra check that the
        field applies at all, so the specialist must be able to tell the two apart.
        """
        engineered = detected = 0
        for name, counts in self.block["per_file"].items():
            if not counts["amber_cells_filled"]:
                continue
            a = self.audit(name)
            for rid in counts["amber_filled_record_ids"]:
                engineered += 1
                hits = [ans for ans in a.by_record.get(rid, [])
                        if ans.field == "Adjuvant therapy given"]
                with self.subTest(name=name, record=rid):
                    self.assertTrue(hits, f"{name}/{rid}: amber answer not reported")
                    self.assertEqual(hits[0].kind, "amber",
                                     "an amber answer must be tagged amber, not yellow")
                detected += 1
        self.assertEqual(engineered, 5, "fixture should fill 5 amber cells")
        self.assertEqual(detected, 5, "all 5 amber answers must be reported")

    def test_gate_context_column_order_is_stable_across_runs(self):
        """The same command on the same data must lay the columns out the same way.

        `build_workbook` used to collect gate-context columns into a SET and insert them at the
        front, so their left-to-right order followed set iteration over strings — which depends
        on PYTHONHASHSEED. Round-to-round worklists weren't diffable, RAs saw columns move, and
        no fixture built through it could be byte-reproducible. `demo_followup` (gates: Sex,
        Status at follow-up) is such a workbook. The order is now the data dictionary's own.
        """
        orders = {}
        for seed in ("0", "2"):
            out = self.out / f"hashseed_{seed}"
            env = dict(os.environ, PYTHONHASHSEED=seed)
            proc = subprocess.run(
                [sys.executable, str(BUILDER),
                 "--records-csv", str(FIXTURE / "records.csv"),
                 "--metadata-csv", str(FIXTURE / "datadictionary.csv"),
                 "--fields", str(FIXTURE / "qa_fields.yaml"),
                 "--out", str(out), "--id-field", "syn_id", "--round="],
                capture_output=True, text=True, timeout=300, env=env)
            self.assertEqual(proc.returncode, 0, proc.stderr[-1000:])
            ws = openpyxl.load_workbook(
                out / "with_MDC" / "demo_followup_site_alpha.xlsx").active
            orders[seed] = [c.value for c in ws[1]]
        self.assertEqual(orders["0"], orders["2"],
                         "gate-context column order must not depend on PYTHONHASHSEED")
        # And the order is the DD's: sex precedes follow_status in datadictionary.csv.
        self.assertEqual(orders["0"][:3], ["syn_id", "Sex", "Status at follow-up"],
                         "gate-context columns come first, in data-dictionary order")

    def test_single_id_column_workbook_recovers_every_answer(self):
        """A workbook with ONE id column must not lose its first data column.

        `_yellow_keys(orig_ws, id_col_count=2)` hardcoded TWO id columns and scanned from
        column 3. `build_worklists.py` writes `--id-field` plus whatever `--extra-id-cols`
        says — one id column by default — so on a default workbook the first DATA column was
        invisible and every RA answer in it was silently discarded.

        The SYN fixture's two workbooks happen to dodge it (gate-context columns sit in front,
        so column 2 is never flagged in either), so this reproduces it directly on a one-field
        workbook: 15 answered cells in, 15 records out.
        """
        cfg = self.out / "one_field.yaml"
        cfg.write_text("workbooks:\n  - name: onefield\n    title: One Field\n"
                       "    fields: [histology_grade]\n")
        build = self.out / "single_id_build"
        proc = subprocess.run(
            [sys.executable, str(BUILDER),
             "--records-csv", str(FIXTURE / "records.csv"),
             "--metadata-csv", str(FIXTURE / "datadictionary.csv"),
             "--fields", str(cfg), "--out", str(build), "--id-field", "syn_id", "--round="],
            capture_output=True, text=True, timeout=300)
        self.assertEqual(proc.returncode, 0, proc.stderr[-1000:])
        src = build / "with_MDC" / "onefield_site_alpha.xlsx"
        wb = openpyxl.load_workbook(src)
        ws = wb.active
        self.assertEqual([c.value for c in ws[1]], ["syn_id", "Histology grade", "RESPONSE"])
        self.assertEqual(self.rr.id_column_count(ws), 1,
                         "the workbook's frozen pane says where the ID block ends")
        answered = 0
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(row=r, column=2)
            if str(cell.fill.fgColor.rgb or "").upper().endswith(YELLOW_HEX):
                cell.value = "Poorly differentiated"
                answered += 1
        dst = build / "onefield_site_alpha_RETURNED.xlsx"
        wb.save(dst)
        self.assertEqual(answered, 15, "fixture should give 15 yellow cells in this workbook")
        a = self.rr.diff(str(src), str(dst))
        self.assertEqual(len(a.by_record), 15, "every answer in the first data column recovered")
        self.assertEqual(sum(len(v) for v in a.by_record.values()), 15)
        self.assertEqual(a.out_of_scope, [], "nothing else was touched")


@unittest.skipIf(not DEPS, "pandas/openpyxl/yaml not installed")
class TestGenerateReturnsFromArbitraryWorklists(unittest.TestCase):
    """`--from-worklists`: returns for whatever a live session actually built.

    The fixture's own two-workbook config is not what the DD-driven skill produces in a real
    session, and the layout mismatch made one Cowork round conclude "the RAs merged the
    workbooks". The generator now takes a worklists directory and derives its edit budget from
    each workbook it finds there.
    """
    @classmethod
    def setUpClass(cls):
        cls.out = Path(tempfile.mkdtemp())
        cls.cfg = cls.out / "live_fields.yaml"
        cls.cfg.write_text(
            "workbooks:\n"
            "  - name: demographics\n    title: Demographics\n"
            "    fields: [sex, age, education, pregnancy_status]\n"
            "  - name: pathology\n    title: Pathology\n"
            "    fields: [histology_grade, margin_status, cea_level]\n")
        cls.worklists = cls.out / "worklists"
        cls.build = subprocess.run(
            [sys.executable, str(BUILDER),
             "--records-csv", str(FIXTURE / "records.csv"),
             "--metadata-csv", str(FIXTURE / "datadictionary.csv"),
             "--fields", str(cls.cfg), "--out", str(cls.worklists), "--id-field", "syn_id"],
            capture_output=True, text=True, timeout=300)
        cls.returns = cls.out / "returns"
        cls.gen = subprocess.run(
            [sys.executable, str(GENERATOR), "--out", str(cls.returns),
             "--from-worklists", str(cls.worklists)],
            capture_output=True, text=True, timeout=600)
        cls.rr = _load_reviewer()

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.out, ignore_errors=True)

    def test_both_steps_succeed(self):
        self.assertEqual(self.build.returncode, 0, self.build.stderr[-1000:])
        self.assertEqual(self.gen.returncode, 0,
                         f"{self.gen.stdout[-1000:]}\n{self.gen.stderr[-1000:]}")
        self.assertNotIn("Traceback", self.gen.stdout + self.gen.stderr)

    def test_a_return_is_written_for_every_worklist(self):
        """It must find the workbooks through the builder's per-round subdir, too."""
        built = sorted(p.stem for p in self.worklists.rglob("with_MDC/*.xlsx"))
        returned = sorted(p.stem[: -len("_RETURNED")]
                          for p in (self.returns / "returned").glob("*_RETURNED.xlsx"))
        self.assertTrue(built, "the live-style build produced no with_MDC worklists")
        self.assertEqual(returned, built)

    def test_the_audit_reports_exactly_what_was_engineered(self):
        """The whole point: returns that match the layout the session built, and an audit that
        reproduces the generator's own numbers on them."""
        counts = json.loads((self.returns / "returned" / "returned_counts.json").read_text())
        self.assertIn("source_dir", counts, "the block should record where it read from")
        for name, spec in counts["expected_review_responses"].items():
            orig = next(self.worklists.rglob(f"with_MDC/{name}.xlsx"))
            resp = self.returns / "returned" / f"{name}_RETURNED.xlsx"
            with self.subTest(name):
                a = self.rr.diff(str(orig), str(resp))
                yellow = sum(1 for cells in a.by_record.values()
                             for ans in cells if ans.kind == "yellow")
                amber = sum(1 for cells in a.by_record.values()
                            for ans in cells if ans.kind == "amber")
                note_only = [rid for rid, n in a.notes.items() if n and rid not in a.by_record]
                self.assertEqual(len(a.by_record), spec["records_with_proposed_updates"])
                self.assertEqual(yellow, spec["changed_cells"])
                self.assertEqual(amber, spec["amber_cells_detected"])
                self.assertEqual(len(note_only), spec["note_only_records"])
                self.assertEqual(len(a.out_of_scope), spec["out_of_scope_edits_detected"])

    def test_update_manifest_is_refused_for_an_arbitrary_directory(self):
        """MANIFEST's `returned` block describes the committed fixture's own build. Letting a
        one-off run overwrite it with counts nothing else can reproduce would poison the suite.
        """
        proc = subprocess.run(
            [sys.executable, str(GENERATOR), "--out", str(self.out / "nope"),
             "--from-worklists", str(self.worklists), "--update-manifest"],
            capture_output=True, text=True, timeout=300)
        self.assertNotEqual(proc.returncode, 0)
        self.assertIn("--update-manifest", proc.stdout + proc.stderr)


@unittest.skipIf(not DEPS, "pandas/openpyxl/yaml not installed")
class TestSummarizeForRAFileMode(unittest.TestCase):
    """The audit's LAST step must be token-optional like every other step.

    `summarize_for_ra.py` used to hard-require --url/--token-env, so a QA round done entirely
    from downloaded files hit a wall at the point of writing the RAs their summaries, and one
    live session hand-wrote them instead. It now takes --metadata-csv, the same Data Dictionary
    file build_worklists.py accepts.
    """
    @classmethod
    def setUpClass(cls):
        cls.out = Path(tempfile.mkdtemp())
        (cls.out / "RA_questions.md").write_text(
            "# Open questions\n"
            "\n## SITE_ALPHA\n"
            '### SYN-0003 — could you clarify your "RESOLVED" note?\n'
            "The cell is still blank in REDCap.\n"
            "\n## SITE_BETA\n"
            '### SYN-0125 — what does "transferred" mean here?\n')
        drafts = cls.out / "push_drafts"
        drafts.mkdir()
        with open(drafts / "sitealpha_clinical.csv", "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["syn_id", "histology_grade"])
            w.writerow(["SYN-0003", "3"])

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.out, ignore_errors=True)

    def run_summarizer(self, *extra):
        return subprocess.run(
            [sys.executable, str(SUMMARIZER),
             "--questions", str(self.out / "RA_questions.md"),
             "--id-field", "syn_id", "--round=", *extra],
            capture_output=True, text=True, timeout=300)

    def test_file_mode_writes_summaries_without_any_key(self):
        out = self.out / "summaries"
        proc = self.run_summarizer(
            "--metadata-csv", str(FIXTURE / "datadictionary.csv"),
            "--push-drafts", str(self.out / "push_drafts"), "--out", str(out))
        self.assertEqual(proc.returncode, 0, f"{proc.stdout[-800:]}\n{proc.stderr[-800:]}")
        self.assertNotIn("Traceback", proc.stdout + proc.stderr)
        self.assertNotIn("access key", proc.stdout + proc.stderr)
        # One file per site named in the questions, plus the push-draft site.
        written = sorted(p.name for p in out.glob("*.md"))
        self.assertIn("site_alpha.md", written)
        self.assertIn("site_beta.md", written)
        alpha = (out / "site_alpha.md").read_text()
        self.assertIn("could you clarify", alpha, "the RA's open question must be copied over")
        # The Data Dictionary was really read: the field code is rendered as its label.
        sitealpha = (out / "sitealpha.md").read_text()
        self.assertIn("Histology grade", sitealpha)

    def test_a_missing_push_drafts_folder_is_fine(self):
        """A normal QA round stages nothing — push_drafts is a migration-only input."""
        out = self.out / "summaries_no_drafts"
        proc = self.run_summarizer(
            "--metadata-csv", str(FIXTURE / "datadictionary.csv"),
            "--push-drafts", str(self.out / "there_is_no_such_folder"), "--out", str(out))
        self.assertEqual(proc.returncode, 0, f"{proc.stdout[-800:]}\n{proc.stderr[-800:]}")
        self.assertTrue((out / "site_alpha.md").exists())

    def test_no_data_source_at_all_explains_both_ways_in_plain_words(self):
        proc = self.run_summarizer("--out", str(self.out / "unused"))
        self.assertNotEqual(proc.returncode, 0)
        msg = proc.stdout + proc.stderr
        self.assertIn("--metadata-csv", msg)
        self.assertIn("--token-env", msg)
        self.assertNotIn("Traceback", msg)


class TestHighlightColoursHaveOneDefinition(unittest.TestCase):
    """0.17.2 #29: "yellow" was #FFC7CE — a pale rose.

    Every instruction to every RA says "fill in the yellow cells", which is a sentence you
    cannot follow when the cells are pink. Worse, the hex was retyped in four files: the builder
    that paints it, the reviewer and the ingester that recognise it coming back, and the fixture
    generator that imitates an RA. Any one of them drifting silently discards a site's answers.
    One definition, imported everywhere.
    """

    FILES = {
        "build_worklists.py": QA_SKILL / "build_worklists.py",
        "review_responses.py": QA_SKILL / "review_responses.py",
        "ingest_response.py": QA_SKILL / "ingest_response.py",
        "generate_returns.py": GENERATOR,
    }

    def test_the_fill_is_actually_yellow(self):
        r, g, b = (int(YELLOW_HEX[i:i + 2], 16) for i in (0, 2, 4))
        self.assertEqual(r, g, f"{YELLOW_HEX} isn't yellow: red and green must match")
        self.assertGreater(r, 200, "yellow has to be bright")
        self.assertLess(b, r - 40, f"{YELLOW_HEX} has too much blue to read as yellow")

    def test_amber_stays_distinguishable(self):
        self.assertNotEqual(AMBER_HEX, YELLOW_HEX)
        self.assertEqual(AMBER_HEX, "FFE9B8", "the amber 'please check' fill is unchanged")

    def test_nobody_retypes_a_colour(self):
        for name, path in self.FILES.items():
            text = path.read_text()
            with self.subTest(name):
                self.assertIn("from qa_colours import", text,
                              f"{name} must import the colours, not define its own")
                self.assertNotIn(f'"{YELLOW_HEX}"', text, f"{name} retypes the yellow hex")
                self.assertNotIn("FFC7CE", text,
                                 f"{name} hardcodes the old rose fill — the reader gets it from "
                                 "qa_colours.LEGACY_FLAG_HEXES and nothing paints it any more")

    def test_the_retired_rose_is_readable_and_defined_once(self):
        """0.19 #42: retired, not forgotten. Returns painted before 0.18 still arrive."""
        import review_responses  # noqa: E402
        self.assertIn("FFC7CE", LEGACY_FLAG_HEXES,
                      "the rose 'yellow' shipped for months; it can never stop being readable")
        self.assertIn("LEGACY_FLAG_HEXES", (QA_SKILL / "review_responses.py").read_text(),
                      "the reader must take the retired colours from qa_colours")
        self.assertEqual(review_responses.LEGACY_FLAG_HEXES, LEGACY_FLAG_HEXES)
        for hexv in LEGACY_FLAG_HEXES:
            self.assertEqual(review_responses.FLAG_KINDS[hexv], "yellow",
                             "a retired flag asked the yellow question, not the amber one")

    @unittest.skipIf(not DEPS, "pandas/openpyxl/yaml not installed")
    def test_a_retired_colour_is_never_a_live_one(self):
        """If a live fill ever landed in the retired tuple the two would be indistinguishable."""
        import build_worklists  # noqa: E402
        painted = {str(build_worklists.YELLOW.start_color.rgb)[-6:],
                   str(build_worklists.UNCERTAIN.start_color.rgb)[-6:]}
        for hexv in LEGACY_FLAG_HEXES:
            self.assertNotIn(hexv, (YELLOW_HEX, AMBER_HEX))
            self.assertNotIn(hexv, painted, "the builder must not paint a retired colour")

    def test_all_four_readers_agree_on_both_colours(self):
        """Loaded, not grepped: the values the running code actually holds."""
        import review_responses, ingest_response, build_worklists  # noqa: E402
        self.assertEqual(review_responses.YELLOW_HEX, YELLOW_HEX)
        self.assertEqual(review_responses.AMBER_HEX, AMBER_HEX)
        self.assertEqual(ingest_response.YELLOW_HEX, YELLOW_HEX)
        self.assertEqual(str(build_worklists.YELLOW.start_color.rgb)[-6:], YELLOW_HEX)
        self.assertEqual(str(build_worklists.UNCERTAIN.start_color.rgb)[-6:], AMBER_HEX)
        spec = importlib.util.spec_from_file_location("_argo_gen_returns", GENERATOR)
        gen = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gen)
        self.assertEqual(gen.YELLOW_HEX, YELLOW_HEX)
        self.assertEqual(gen.AMBER_HEX, AMBER_HEX)

    @unittest.skipIf(not DEPS, "pandas/openpyxl/yaml not installed")
    def test_a_built_workbook_really_is_painted_that_colour(self):
        """Not just the constant — the cells an RA opens."""
        out = Path(tempfile.mkdtemp())
        try:
            cfg = out / "one.yaml"
            cfg.write_text("workbooks:\n  - name: onefield\n    title: One\n"
                           "    fields: [histology_grade]\n")
            proc = subprocess.run(
                [sys.executable, str(BUILDER),
                 "--records-csv", str(FIXTURE / "records.csv"),
                 "--metadata-csv", str(FIXTURE / "datadictionary.csv"),
                 "--fields", str(cfg), "--out", str(out / "b"),
                 "--id-field", "syn_id", "--round="],
                capture_output=True, text=True, timeout=300)
            self.assertEqual(proc.returncode, 0, proc.stderr[-800:])
            ws = openpyxl.load_workbook(out / "b" / "with_MDC" /
                                        "onefield_site_alpha.xlsx").active
            fills = {str(ws.cell(row=r, column=2).fill.fgColor.rgb or "").upper()[-6:]
                     for r in range(3, ws.max_row + 1)}
            self.assertIn(YELLOW_HEX, fills, "no cell was painted the yellow the RA is told about")
            self.assertNotIn("FFC7CE", fills)
        finally:
            shutil.rmtree(out, ignore_errors=True)


@unittest.skipIf(not DEPS, "pandas/openpyxl/yaml not installed")
class TestLegacyRoseWorklistsStillAudit(unittest.TestCase):
    """0.19 #42: a worklist painted the OLD rose must audit exactly like a yellow one.

    "Yellow" was `FFC7CE` — a pale rose — until 0.18 made it an actual yellow. Worklists don't
    come back the day they're sent: a site that received one before the change returns it weeks
    later, still rose, with the work done. The reader matched only the current yellow, so a live
    round reported 5 of 36 answers and said nothing whatever about the other 31 — the exact
    silent-data-loss shape this suite exists to catch.

    The same build, the same RA edits, painted two ways. The audits must be identical, and the
    rose one must say ONE line about why it's reading a colour nothing paints any more.
    """

    ROSE = LEGACY_FLAG_HEXES[0]

    @classmethod
    def setUpClass(cls):
        cls.tmp = Path(tempfile.mkdtemp())
        cfg = cls.tmp / "one.yaml"
        cfg.write_text("workbooks:\n  - name: onefield\n    title: One Field\n"
                       "    fields: [histology_grade]\n")
        proc = subprocess.run(
            [sys.executable, str(BUILDER),
             "--records-csv", str(FIXTURE / "records.csv"),
             "--metadata-csv", str(FIXTURE / "datadictionary.csv"),
             "--fields", str(cfg), "--out", str(cls.tmp / "build"),
             "--id-field", "syn_id", "--round="],
            capture_output=True, text=True, timeout=300)
        if proc.returncode != 0:                       # pragma: no cover - setup failure
            raise RuntimeError(proc.stderr[-1000:])
        cls.modern = cls.tmp / "build" / "with_MDC" / "onefield_site_alpha.xlsx"
        cls.rr = _load_reviewer()
        cls.legacy, cls.filled = cls._repaint_and_answer()

    @classmethod
    def _repaint_and_answer(cls):
        """A rose copy of the original, plus a return for each — same answers in both."""
        from openpyxl.styles import PatternFill
        rose = PatternFill(start_color=cls.ROSE, end_color=cls.ROSE, fill_type="solid")

        legacy_orig = cls.tmp / "legacy_original.xlsx"
        wb = openpyxl.load_workbook(cls.modern)
        ws = wb.active
        painted = []
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(row=r, column=2)
            if str(cell.fill.fgColor.rgb or "").upper().endswith(YELLOW_HEX):
                cell.fill = rose
                painted.append(r)
        wb.save(legacy_orig)

        for src, dst in ((cls.modern, cls.tmp / "modern_RETURNED.xlsx"),
                         (legacy_orig, cls.tmp / "legacy_RETURNED.xlsx")):
            wb = openpyxl.load_workbook(src)
            ws = wb.active
            for r in painted:
                ws.cell(row=r, column=2).value = "Poorly differentiated"
                ws.cell(row=r, column=ws.max_column).value = "Entered in REDCap."
            wb.save(dst)
        return legacy_orig, painted

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmp, ignore_errors=True)

    def test_the_fixture_really_is_painted_two_different_colours(self):
        """Guard the guard: if both files were yellow this test would prove nothing."""
        self.assertTrue(self.filled, "no cell was flagged, so there is nothing to audit")
        for path, hexv in ((self.modern, YELLOW_HEX), (self.legacy, self.ROSE)):
            ws = openpyxl.load_workbook(path).active
            fills = {str(ws.cell(row=r, column=2).fill.fgColor.rgb or "").upper()[-6:]
                     for r in self.filled}
            self.assertEqual(fills, {hexv}, path.name)

    def test_the_rose_workbook_audits_to_the_same_counts_as_the_yellow_one(self):
        modern = self.rr.diff(str(self.modern), str(self.tmp / "modern_RETURNED.xlsx"))
        legacy = self.rr.diff(str(self.legacy), str(self.tmp / "legacy_RETURNED.xlsx"))
        self.assertEqual(len(legacy.by_record), len(modern.by_record))
        self.assertEqual(sum(len(v) for v in legacy.by_record.values()),
                         sum(len(v) for v in modern.by_record.values()))
        self.assertEqual(len(self.filled), sum(len(v) for v in legacy.by_record.values()),
                         "every answer in the rose workbook must be recovered")
        self.assertEqual(legacy.out_of_scope, modern.out_of_scope, [])
        self.assertEqual(sorted(legacy.notes), sorted(modern.notes))

    def test_a_rose_flag_is_reported_as_yellow_not_amber(self):
        """The old rose MEANT "this applies and is blank" — the yellow question, not the amber
        one. Reporting it as amber would tell the specialist to go and check a condition that
        was never in doubt."""
        legacy = self.rr.diff(str(self.legacy), str(self.tmp / "legacy_RETURNED.xlsx"))
        kinds = {a.kind for cells in legacy.by_record.values() for a in cells}
        self.assertEqual(kinds, {"yellow"})

    def test_one_warning_line_when_the_flags_are_mostly_legacy(self):
        proc = subprocess.run(
            [sys.executable, str(REVIEWER), str(self.legacy),
             str(self.tmp / "legacy_RETURNED.xlsx")],
            capture_output=True, text=True, timeout=300)
        self.assertEqual(proc.returncode, 0, proc.stderr[-800:])
        hits = [ln for ln in proc.stdout.splitlines() if "older version" in ln]
        self.assertEqual(len(hits), 1, f"expected exactly one warning line, got {hits}")
        self.assertIn("reading the old", hits[0].lower())

    def test_a_normal_workbook_says_nothing_about_old_colours(self):
        proc = subprocess.run(
            [sys.executable, str(REVIEWER), str(self.modern),
             str(self.tmp / "modern_RETURNED.xlsx")],
            capture_output=True, text=True, timeout=300)
        self.assertEqual(proc.returncode, 0, proc.stderr[-800:])
        self.assertNotIn("older version", proc.stdout)

    def test_a_single_recoloured_cell_is_not_worth_a_warning(self):
        """Somebody hand-shading one cell isn't an old build, and shouldn't read like one."""
        note = self.rr.legacy_flag_note
        self.assertEqual(note(20, 1), "")
        self.assertEqual(note(20, 10), "", "exactly half is not 'predominantly'")
        self.assertIn("older version", note(20, 11))
        self.assertEqual(note(0, 0), "", "a workbook with no flags at all says nothing")


class TestSiteHeadersAreWholeHeaders(unittest.TestCase):
    """0.17.2 #32: `parse_questions` keyed on the header's FIRST WORD.

    `## Site Alpha` and `## Site Beta` both became "site", so one site's open questions were
    served to every RA in the study. The key is now the whole header, lowercased with its
    whitespace collapsed — and two headers that collapse to the same key are merged loudly,
    because the quiet reading of that situation is one RA receiving another site's questions.
    """

    @classmethod
    def setUpClass(cls):
        spec = importlib.util.spec_from_file_location("_argo_summarize", SUMMARIZER)
        cls.mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(cls.mod)

    def _questions(self, text, warn=lambda *_a: None):
        tmp = Path(tempfile.mkdtemp())
        try:
            path = tmp / "RA_questions.md"
            path.write_text(text)
            return self.mod.parse_questions(str(path), warn=warn)
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    DOC = ("# Open questions\n"
           "\n## Site Alpha\n### SYN-0003 — alpha's question\n\n"
           "## Site Beta\n### SYN-0125 — beta's question\n")

    def test_two_multiword_sites_stay_apart(self):
        got = self._questions(self.DOC)
        self.assertEqual(sorted(got), ["site alpha", "site beta"])
        self.assertIn("alpha's question", got["site alpha"])
        self.assertNotIn("beta's question", got["site alpha"])

    def test_the_first_word_is_not_the_key(self):
        """The exact defect: both sites used to collapse onto 'site'."""
        self.assertNotIn("site", self._questions(self.DOC))

    def test_case_and_spacing_are_ignored(self):
        got = self._questions("## SITE   ALPHA\n### q\n")
        self.assertEqual(list(got), ["site alpha"])

    def test_a_genuine_collision_is_merged_but_warned_about(self):
        warnings = []
        got = self._questions(
            "## Site Alpha\n### first\n\n## site  alpha\n### second\n", warn=warnings.append)
        self.assertEqual(list(got), ["site alpha"])
        self.assertIn("first", got["site alpha"])
        self.assertIn("second", got["site alpha"])
        self.assertEqual(len(warnings), 1, warnings)
        self.assertIn("same site", warnings[0])

    def test_an_underscored_site_name_is_unchanged(self):
        """`## SITE_ALPHA` was already a single word, and must keep behaving identically."""
        self.assertEqual(list(self._questions("## SITE_ALPHA\n### q\n")), ["site_alpha"])


if __name__ == "__main__":
    unittest.main(verbosity=2)
