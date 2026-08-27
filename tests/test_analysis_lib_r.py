#!/usr/bin/env python3
"""The R half of the ARGO analysis library, checked against the golden Table 1.

`plugins/argo-data-analyst/skills/run-analysis/lib/R/argo_analysis/` is a
SOURCED library (not a package): core.R, table1.R, excel.R, figures.R,
survival.R, plus a one-command runner, run_table1.R. It is the mirror of the
Python library in lib/python/argo_analysis/ -- same function names, same table
shape, same rounding.

What is checked here, all against the committed synthetic fixture:

    * the runner reproduces testing/fixtures/synthetic-study/analysis/
      expected_table1.csv -- every number in the golden, to within 0.01
    * the applicable denominator: pregnancy_status is out of 111 women, not
      out of 200 participants
    * branching logic we cannot read WARNS and counts everyone, never silently
      drops the field
    * the survival module stops with the planned message
    * write_workbook() produces a workbook that opens, with the sheets and the
      Notes sheet the house style requires
    * the figures are real PNGs

R is found through the toolkit's own preflight (argo_tools.detect), which probes
the folders a Cowork shell's thin PATH misses, and which reports whether Rscript
can actually RUN something rather than only whether the file exists. Where R is
not installed, or is installed but broken, every R test here skips cleanly --
the suite must pass on a machine with no R at all.

No network, no keys, no patient data.
"""
from __future__ import annotations

import csv
import importlib.util
import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
FIXTURE = REPO / "testing" / "fixtures" / "synthetic-study"
GOLDEN = FIXTURE / "analysis" / "expected_table1.csv"
SKILL_DIR = REPO / "plugins" / "argo-data-analyst" / "skills" / "run-analysis"
LIB = SKILL_DIR / "lib" / "R" / "argo_analysis"
RUNNER = LIB / "run_table1.R"
SCAFFOLD = SKILL_DIR / "scaffold.py"

# The variables the golden describes, in the golden's order.
GOLDEN_VARIABLES = "age,sex,education,marital_status,tobacco_use,histology_grade"

# R and Python both round through C's printf, so in practice they agree to the
# last printed digit. The tolerance is there so a platform whose printf breaks a
# rounding tie the other way fails on the tie alone, not on the whole table.
TOL = 0.01

KEY = ("variable", "level", "statistic")


def _find_rscript():
    """R via the toolkit's preflight: is it there, and can it actually run?"""
    tools = REPO / "plugins/argo-core/skills/redcap-api/scripts/argo_tools.py"
    try:
        spec = importlib.util.spec_from_file_location("argo_tools", tools)
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        info = mod.detect(probe_versions=False, verify_run=True).get("r") or {}
    except Exception:
        path = shutil.which("Rscript")
        return path, ("Rscript is not on PATH" if path is None else None)
    if not info.get("found"):
        return None, "R is not installed on this computer"
    if info.get("runs") is False:
        return None, f"R is installed but will not run: {info.get('run_error')}"
    return info.get("path") or shutil.which("Rscript"), None


RSCRIPT, R_SKIP_REASON = _find_rscript()
SKIP = R_SKIP_REASON or "R not available"


def read_table(path):
    with open(path, newline="") as fh:
        reader = csv.DictReader(fh)
        return reader.fieldnames, list(reader)


def run_r(script_source, cwd):
    """Run a snippet of R that has already sourced whatever it needs."""
    script = Path(cwd) / "snippet.R"
    script.write_text(script_source)
    return subprocess.run([RSCRIPT, str(script)], capture_output=True, text=True,
                          timeout=300, cwd=str(cwd))


def source_lines():
    """The `source()` calls that put the whole library in scope."""
    return "\n".join(f'source({str(LIB / name)!r})'
                     for name in ("core.R", "table1.R", "excel.R", "figures.R",
                                  "survival.R")) + "\n"


def load_fixture_lines():
    return (f'study <- apply_missing(load_study({str(FIXTURE / "records.csv")!r}, '
            f'{str(FIXTURE / "datadictionary.csv")!r}))\n')


class TestTheLibraryIsPresent(unittest.TestCase):
    """Runs everywhere, R or no R: the files have to exist to be shipped."""

    def test_every_module_of_the_contract_is_there(self):
        for name in ("core.R", "table1.R", "excel.R", "figures.R", "survival.R",
                     "run_table1.R"):
            self.assertTrue((LIB / name).is_file(), f"lib/R/argo_analysis/{name} is missing")

    def test_the_contract_functions_are_defined(self):
        expected = {
            "core.R": ("load_study", "apply_missing", "labels", "applicable", "denominator"),
            "table1.R": ("table1",),
            "excel.R": ("write_workbook",),
            "figures.R": ("bar_by_group", "hist"),
            "survival.R": ("survival",),
        }
        for module, functions in expected.items():
            text = (LIB / module).read_text()
            for fn in functions:
                self.assertIn(f"{fn} <- function", text,
                              f"{module} does not define {fn}()")

    def test_survival_is_marked_planned_in_the_source(self):
        text = (LIB / "survival.R").read_text()
        self.assertIn("Survival analysis is planned but not built yet", text)

    def test_the_legend_never_goes_back_inside_the_plot(self):
        """A source guard, so it holds on a machine with no R: both figures draw their
        key through the one helper that puts it in the reserved right margin."""
        text = (LIB / "figures.R").read_text()
        self.assertIn("argo_legend_right <- function", text,
                      "the outside-the-plot legend helper is gone")
        self.assertNotIn('legend("topright"', text,
                         "a legend at topright covers the tallest bar")
        bar = text.split("bar_by_group <- function", 1)[1].split("\nhist <- function", 1)[0]
        histogram = text.split("\nhist <- function", 1)[1]
        for body, name in ((bar, "bar_by_group"), (histogram, "hist")):
            self.assertIn("argo_legend_right(", body,
                          f"{name} does not use the outside-the-plot legend")
            self.assertIn("argo_legend_margin_lines(", body,
                          f"{name} draws before reserving the margin the legend needs")

    def test_statistics_are_converted_to_numbers_before_they_are_written(self):
        """A source guard for the same rule the workbook test measures: excel.R must
        hand openxlsx numbers, and must leave the label columns alone."""
        text = (LIB / "excel.R").read_text()
        self.assertIn("argo_numeric_cells <- function", text)
        self.assertIn("argo_numeric_cells(tb)", text,
                      "write_workbook does not convert its tables before writing")
        for label in ("variable", "level", "statistic"):
            self.assertIn(f'"{label}"', text.split("ARGO_LABEL_COLUMNS", 1)[1][:200],
                          f"{label} is not protected from being turned into a number")

    def test_the_only_package_dependency_is_guarded(self):
        """openxlsx is the one add-on the library uses, and a missing add-on must
        produce an install line a non-programmer can copy, not a stack trace."""
        text = (LIB / "excel.R").read_text()
        self.assertIn('requireNamespace("openxlsx"', text.replace("ARGO_EXCEL_PACKAGE", '"openxlsx"'))
        self.assertIn('install.packages("openxlsx")', text)
        for module in ("core.R", "table1.R", "figures.R", "survival.R"):
            body = (LIB / module).read_text()
            self.assertNotIn("library(", body,
                             f"{module} must be base R only -- no library() calls")


@unittest.skipIf(RSCRIPT is None, SKIP)
class TestRunnerReproducesTheGolden(unittest.TestCase):
    """run_table1.R, one command, against the committed golden Table 1."""

    @classmethod
    def setUpClass(cls):
        cls.tmp = Path(tempfile.mkdtemp())
        cls.out = cls.tmp / "table1.csv"
        cls.proc = subprocess.run(
            [RSCRIPT, str(RUNNER),
             "--export", str(FIXTURE / "records.csv"),
             "--dictionary", str(FIXTURE / "datadictionary.csv"),
             "--group-by", "redcap_data_access_group",
             "--variables", GOLDEN_VARIABLES,
             "--out", str(cls.out)],
            capture_output=True, text=True, timeout=300, cwd=str(cls.tmp))

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmp, ignore_errors=True)

    def test_the_runner_runs(self):
        self.assertEqual(self.proc.returncode, 0,
                         f"run_table1.R failed:\n{self.proc.stderr[-2000:]}")
        self.assertTrue(self.out.exists(), "run_table1.R wrote no table")

    def test_the_columns_are_the_goldens_columns(self):
        header, _ = read_table(self.out)
        g_header, _ = read_table(GOLDEN)
        self.assertEqual(header, g_header)

    def test_every_number_in_the_golden_is_reproduced(self):
        g_header, g_rows = read_table(GOLDEN)
        _, r_rows = read_table(self.out)
        produced = {tuple(r[k] for k in KEY): r for r in r_rows}
        value_columns = g_header[3:]
        for g in g_rows:
            key = tuple(g[k] for k in KEY)
            self.assertIn(key, produced, f"the R table is missing the row {key}")
            r = produced[key]
            for col in value_columns:
                gv, rv = g[col].strip(), r[col].strip()
                if gv == "" or rv == "":
                    self.assertEqual(gv, rv, f"{key}/{col}: one side blank")
                    continue
                self.assertAlmostEqual(float(gv), float(rv), delta=TOL,
                                       msg=f"{key}/{col}: golden={gv} R={rv}")

    def test_the_golden_rows_keep_the_goldens_order(self):
        """Level order comes from the codebook, so the shared rows must appear in
        the same order -- an alphabetised Table 1 is a different table."""
        _, g_rows = read_table(GOLDEN)
        _, r_rows = read_table(self.out)
        wanted = [tuple(r[k] for k in KEY) for r in g_rows]
        got = [tuple(r[k] for k in KEY) for r in r_rows]
        self.assertEqual([k for k in got if k in set(wanted)], wanted)

    def test_the_contracts_extra_continuous_statistics_are_there(self):
        """The golden predates median/q1/q3; the contract requires them, so they
        are extra rows rather than changed ones."""
        _, r_rows = read_table(self.out)
        stats = {r["statistic"] for r in r_rows if r["variable"] == "age"}
        self.assertEqual(stats, {"n", "missing", "mean", "sd", "median", "q1", "q3"})


@unittest.skipIf(RSCRIPT is None, SKIP)
class TestApplicableDenominator(unittest.TestCase):
    """The rule the whole library turns on: describe a field against the records
    it was actually asked of."""

    @classmethod
    def setUpClass(cls):
        cls.tmp = Path(tempfile.mkdtemp())

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmp, ignore_errors=True)

    def r(self, body):
        proc = run_r(source_lines() + load_fixture_lines() + body, self.tmp)
        self.assertEqual(proc.returncode, 0, f"R failed:\n{proc.stderr[-2000:]}")
        return proc

    def test_pregnancy_status_is_out_of_the_111_women(self):
        """[sex] = '2'. 200 records, 111 of them female. The naive denominator,
        200, is the bug this rule exists to prevent."""
        out = self.r('cat(denominator(study, "pregnancy_status"))').stdout.strip()
        self.assertEqual(out, "111")

    def test_a_field_with_no_branching_applies_to_everyone(self):
        self.assertEqual(self.r('cat(denominator(study, "sex"))').stdout.strip(), "200")

    def test_the_supported_branching_grammar(self):
        """One denominator per shape the fixture's MANIFEST engineers: quoted,
        unquoted, checkbox, numeric comparison, AND, OR."""
        body = ('for (f in c("pregnancy_status", "dx_date", "bleeding_severity", '
                '"alcohol_use", "chemo_cycles", "support_needed")) '
                'cat(f, denominator(study, f), "\\n")')
        counts = dict(line.split() for line in self.r(body).stdout.strip().splitlines())
        self.assertEqual(counts["pregnancy_status"], "111")   # [sex] = '2'      quoted
        self.assertEqual(counts["alcohol_use"], "200")        # [age] >= 18      numeric
        for field in ("dx_date", "bleeding_severity", "chemo_cycles", "support_needed"):
            n = int(counts[field])
            self.assertTrue(0 < n < 200,
                            f"{field}: branching logic did not narrow anything ({n})")

    def test_a_branching_field_is_described_against_its_own_denominator(self):
        """The pregnancy percentages must add to 100 over the women, and the
        counts must add to the applicable denominator -- not to 200."""
        body = ('t <- table1(study, group_by = "redcap_data_access_group", '
                'variables = "pregnancy_status")\n'
                'p <- t[t$variable == "pregnancy_status", ]\n'
                'cat(sum(as.numeric(p$overall[p$statistic == "n"])), '
                'as.numeric(p$overall[p$statistic == "missing"]), '
                'sum(as.numeric(p$overall[p$statistic == "pct"])))')
        counted, missing, pct_total = self.r(body).stdout.strip().split()
        self.assertEqual(int(counted) + int(missing), 111,
                         "pregnancy_status was not described against the 111 women")
        self.assertAlmostEqual(float(pct_total), 100.0, delta=0.05)

    def test_unreadable_branching_warns_and_keeps_every_record(self):
        """datediff(...) is outside the grammar. The field must be counted for
        everyone AND say so -- silently dropping it is the failure this replaced."""
        proc = self.r('cat("N=", denominator(study, "adjuvant_therapy"), "\\n", sep = "")')
        self.assertIn("N=200", proc.stdout)
        warned = proc.stderr + proc.stdout
        self.assertIn("could not fully read the branching condition", warned.lower(),
                      "an unreadable condition was used without warning anyone")
        self.assertIn("datediff", warned,
                      "the warning must quote the condition it could not read")


@unittest.skipIf(RSCRIPT is None, SKIP)
class TestSurvivalIsPlanned(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.tmp = Path(tempfile.mkdtemp())

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmp, ignore_errors=True)

    def test_calling_it_stops_with_the_planned_message(self):
        for name in ("survival", "survival_analysis"):
            proc = run_r(f'source({str(LIB / "survival.R")!r})\n{name}()\n', self.tmp)
            self.assertNotEqual(proc.returncode, 0, f"{name}() did not stop")
            self.assertIn("Survival analysis is planned but not built yet", proc.stderr)

    def test_the_status_the_registry_reads_says_planned(self):
        proc = run_r(f'source({str(LIB / "survival.R")!r})\ncat(ARGO_SURVIVAL_STATUS)\n',
                     self.tmp)
        self.assertEqual(proc.stdout.strip(), "planned")

    def test_it_says_what_does_work(self):
        """A dead end that does not point anywhere is a dead end someone reports
        as a bug."""
        proc = run_r(f'source({str(LIB / "survival.R")!r})\nsurvival()\n', self.tmp)
        self.assertIn("Table 1", proc.stderr)


@unittest.skipIf(RSCRIPT is None, SKIP)
class TestTable1RefusesToGuess(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.tmp = Path(tempfile.mkdtemp())

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmp, ignore_errors=True)

    def test_group_by_is_required(self):
        proc = run_r(source_lines() + load_fixture_lines()
                     + 'table1(study, variables = "sex")\n', self.tmp)
        self.assertNotEqual(proc.returncode, 0,
                            "table1() grouped by something it was never told")
        self.assertIn("group", proc.stderr.lower())

    def test_an_unknown_grouping_variable_is_named_plainly(self):
        proc = run_r(source_lines() + load_fixture_lines()
                     + 'table1(study, group_by = "site", variables = "sex")\n', self.tmp)
        self.assertNotEqual(proc.returncode, 0)
        self.assertIn("site", proc.stderr)
        self.assertNotIn("Error in", proc.stderr.split("\n")[0])


@unittest.skipIf(RSCRIPT is None, SKIP)
class TestWorkbookAndFigures(unittest.TestCase):
    """The house style, checked from the outside: a workbook another tool can
    open, and figures that are really PNGs."""

    @classmethod
    def setUpClass(cls):
        cls.tmp = Path(tempfile.mkdtemp())
        cls.xlsx = cls.tmp / "analysis.xlsx"
        cls.bar = cls.tmp / "figs" / "education.png"
        cls.histogram = cls.tmp / "figs" / "age.png"
        body = (
            't1 <- table1(study, group_by = "redcap_data_access_group",\n'
            '             variables = c("age", "sex", "education"))\n'
            f'write_workbook(list("Table 1" = t1), {str(cls.xlsx)!r},\n'
            '               notes = "Cohort: the whole synthetic export.")\n'
            f'bar_by_group(study, "education", "redcap_data_access_group", {str(cls.bar)!r})\n'
            f'hist(study, "age", {str(cls.histogram)!r})\n'
        )
        cls.proc = run_r(source_lines() + load_fixture_lines() + body, cls.tmp)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmp, ignore_errors=True)

    def _skip_if_openxlsx_missing(self):
        if "install.packages" in self.proc.stderr and not self.xlsx.exists():
            self.skipTest("the openxlsx add-on is not installed in this R")

    def test_the_workbook_was_written(self):
        self._skip_if_openxlsx_missing()
        self.assertEqual(self.proc.returncode, 0,
                         f"the R script failed:\n{self.proc.stderr[-2000:]}")
        self.assertTrue(self.xlsx.exists() and self.xlsx.stat().st_size > 0)

    def test_the_workbook_opens_and_carries_a_notes_sheet(self):
        self._skip_if_openxlsx_missing()
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed, so the workbook cannot be opened here")
        wb = openpyxl.load_workbook(self.xlsx)
        self.assertIn("Table 1", wb.sheetnames)
        self.assertEqual(wb.sheetnames[-1], "Notes", "the Notes sheet must come last")
        notes = " ".join(str(row[0]) for row in wb["Notes"].iter_rows(values_only=True)
                         if row[0])
        self.assertIn("N = 200", notes, "the Notes sheet must state N")
        self.assertIn("-666", notes, "the Notes sheet must state the missing-data rule")
        self.assertIn("applicable denominator", notes,
                      "the Notes sheet must state the denominator rule")
        self.assertIn("Generated by", notes,
                      "the Notes sheet must name the script and the date")
        sheet = wb["Table 1"]
        self.assertTrue(sheet["A1"].font.bold, "the header row must be bold")
        self.assertEqual(sheet.freeze_panes, "A2", "the header row must be frozen")

    def test_the_figures_are_real_pngs(self):
        if self.proc.returncode != 0 and not self.bar.exists():
            self._skip_if_openxlsx_missing()
        for path in (self.bar, self.histogram):
            self.assertTrue(path.exists(), f"{path.name} was not written")
            self.assertGreater(path.stat().st_size, 1000, f"{path.name} is suspiciously small")
            with open(path, "rb") as fh:
                self.assertEqual(fh.read(8), b"\x89PNG\r\n\x1a\n",
                                 f"{path.name} is not a PNG")

    def test_statistics_are_written_as_numbers_not_text(self):
        """openxlsx was handed "2.50" as a string, so the R workbook and the Python
        workbook of the same table were different objects to Excel: text cannot be
        summed, sorted or charted, and a column of it sorts 10 before 9. Labels and
        levels stay text — a level coded "1" is still a label."""
        self._skip_if_openxlsx_missing()
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed, so the workbook cannot be opened here")
        sheet = openpyxl.load_workbook(self.xlsx)["Table 1"]
        header = [cell.value for cell in sheet[1]]
        rows = {(row[0].value, row[2].value): row for row in sheet.iter_rows(min_row=2)}
        columns = [header.index(name) for name in ("site_alpha", "site_beta", "overall")]

        counts = rows[("records", "n")]
        self.assertEqual([counts[i].value for i in columns], [120, 80, 200])
        for i in columns:
            self.assertEqual(counts[i].data_type, "n",
                             f"{header[i]} holds a count as {counts[i].data_type!r}, not a "
                             "number — the Python workbook stores it as one")

        golden = {(r["variable"], r["statistic"]): r for r in read_table(GOLDEN)[1]}
        mean = rows[("age", "mean")]
        for i in columns:
            self.assertEqual(mean[i].data_type, "n", f"{header[i]} holds a mean as text")
            self.assertIsInstance(mean[i].value, (int, float),
                                  f"{header[i]} came back as {type(mean[i].value).__name__}")
            self.assertAlmostEqual(float(mean[i].value),
                                   float(golden[("age", "mean")][header[i]]), delta=TOL,
                                   msg=f"the number itself changed in {header[i]}")
        for name in ("variable", "level", "statistic"):
            cell = mean[header.index(name)]
            self.assertEqual(cell.data_type, "s",
                             f"the {name} column must stay text, not become a number")


@unittest.skipIf(RSCRIPT is None, SKIP)
class TestTheLegendIsOutsideThePlotArea(unittest.TestCase):
    """The legend was drawn at "topright" — inside the plot, on top of the tallest
    bar, which is the bar the reader came for. Measured on a real device: the box the
    legend draws must start at or past the right-hand edge of the axes."""

    def test_the_legend_box_starts_past_the_right_edge_of_the_axes(self):
        tmp = Path(tempfile.mkdtemp())
        try:
            probe = tmp / "probe.png"
            body = (
                'labels <- c("No formal schooling", "Secondary")\n'
                f'argo_open_png({str(probe)!r})\n'
                'graphics::par(mar = c(5, 5, 4, argo_legend_margin_lines(labels)))\n'
                'graphics::barplot(c(80, 20), ylim = c(0, 100))\n'
                'box <- argo_legend_right(labels, fill = c("#0072B2", "#E69F00"))\n'
                'usr <- graphics::par("usr")\n'
                'cat(sprintf("MEASURED %f %f %f\\n", box$rect$left, usr[2], box$rect$w))\n'
                'invisible(grDevices::dev.off())\n'
            )
            proc = run_r(source_lines() + body, tmp)
            self.assertEqual(proc.returncode, 0, proc.stderr[-1500:])
            reported = [ln for ln in proc.stdout.splitlines() if ln.startswith("MEASURED")]
            self.assertTrue(reported, f"the probe measured nothing:\n{proc.stdout}\n"
                                      f"{proc.stderr[-800:]}")
            left, axes_right, width = (float(x) for x in reported[0].split()[1:])
            self.assertGreater(width, 0, "the legend drew nothing")
            self.assertGreaterEqual(left, axes_right,
                                    "the legend starts inside the plot area — it must sit "
                                    "clear of the bars, in the reserved right margin")
            self.assertTrue(probe.is_file() and probe.stat().st_size > 1000,
                            "the probe figure was not written")
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


@unittest.skipIf(RSCRIPT is None, SKIP)
class TestTheScaffoldedRTwinRuns(unittest.TestCase):
    """scaffold.py writes scripts/01_table1.R beside scripts/01_table1.py. What makes
    it a twin rather than a decoration is that it RUNS — against the study folder's own
    copied library, landing the same files the Python one lands."""

    @classmethod
    def setUpClass(cls):
        cls.tmp = Path(tempfile.mkdtemp())
        cls.root = cls.tmp / "study"
        cls.scaffold = subprocess.run(
            [sys.executable, str(SCAFFOLD), str(cls.root),
             "--export", str(FIXTURE / "records.csv"),
             "--dictionary", str(FIXTURE / "datadictionary.csv"),
             "--group-by", "redcap_data_access_group"],
            capture_output=True, text=True, timeout=300)
        cls.script = cls.root / "scripts" / "01_table1.R"
        cls.xlsx = cls.root / "outputs" / "tables" / "table1.xlsx"
        cls.proc = (subprocess.run([RSCRIPT, str(cls.script)], capture_output=True,
                                   text=True, timeout=600, cwd=str(cls.root))
                    if cls.script.is_file() else None)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmp, ignore_errors=True)

    def _skip_if_openxlsx_missing(self):
        if self.proc and "install.packages" in self.proc.stderr and not self.xlsx.exists():
            self.skipTest("the openxlsx add-on is not installed in this R")

    def test_the_scaffold_wrote_it(self):
        self.assertEqual(self.scaffold.returncode, 0, self.scaffold.stderr[-800:])
        self.assertTrue(self.script.is_file(), "scaffold.py wrote no scripts/01_table1.R")

    def test_it_runs_against_the_study_folders_own_library(self):
        self._skip_if_openxlsx_missing()
        self.assertEqual(self.proc.returncode, 0,
                         (self.proc.stdout + self.proc.stderr)[-1500:])
        self.assertNotIn("could not find function", self.proc.stderr,
                         "the generated script calls something the library does not define")

    def test_it_writes_the_same_files_the_python_twin_writes(self):
        self._skip_if_openxlsx_missing()
        self.assertTrue(self.xlsx.is_file(),
                        "no outputs/tables/table1.xlsx — the twins must write the same file")
        figures = sorted((self.root / "outputs" / "figures").glob("*.png"))
        self.assertTrue(figures, "the R twin wrote no figure")
        self.assertTrue(
            any("_by_redcap_data_access_group.png" in f.name for f in figures),
            f"the figure is not named the way the Python twin names it: "
            f"{[f.name for f in figures]}")
        for figure in figures:
            with open(figure, "rb") as fh:
                self.assertEqual(fh.read(8), b"\x89PNG\r\n\x1a\n")

    def test_the_workbook_it_writes_is_the_house_style(self):
        self._skip_if_openxlsx_missing()
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed, so the workbook cannot be opened here")
        book = openpyxl.load_workbook(self.xlsx)
        self.assertIn("Table 1", book.sheetnames)
        self.assertEqual(book.sheetnames[-1], "Notes")
        notes = " ".join(str(row[0]) for row in book["Notes"].iter_rows(values_only=True)
                         if row[0])
        self.assertIn("01_table1.R", notes,
                      "the Notes sheet must name the script that made the file")
        sheet = book["Table 1"]
        row = next(r for r in sheet.iter_rows(min_row=2) if r[0].value == "records")
        numbers = [c for c in row[3:] if c.value is not None]
        self.assertTrue(numbers, "the records row is empty")
        for cell in numbers:
            self.assertEqual(cell.data_type, "n",
                             "the scaffolded R script must produce the same cell types "
                             "as the Python one")


if __name__ == "__main__":
    unittest.main(verbosity=2)
