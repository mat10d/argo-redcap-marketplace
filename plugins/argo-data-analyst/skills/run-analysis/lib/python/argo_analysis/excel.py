#!/usr/bin/env python3
"""argo_analysis.excel — one workbook per analysis, always the same shape.

    from argo_analysis import excel
    excel.write_workbook({"Table 1": t1}, "outputs/table1.xlsx",
                         notes=["Grouped by site."])

THE HOUSE STYLE
---------------
Every workbook this library writes looks the same, so a reader who has opened
one has opened all of them:

  * one workbook per analysis, one sheet per table, in the order given
  * the header row is bold and stays put when you scroll (the top row is frozen)
  * columns are as wide as their contents, so nothing shows as ####
  * numbers are written as numbers, not as text, so they can be re-used
  * the last sheet is always **Notes**, and it carries the four things a table
    is useless without: how many records, what counted as missing, what the
    percentages are out of, and which script produced it, on what date

The Notes sheet is not decoration. A Table 1 emailed without its denominator
rule is a table that will be misread, and the person who misreads it will be
reading it six months after the analyst who made it has moved on.
"""

from __future__ import annotations

import datetime
import re
import sys
from pathlib import Path

try:
    from . import core
except ImportError:                   # run as a loose file from its own folder
    import core                       # type: ignore


#: Standing notes, on every workbook, because they are true of every table here.
STANDING_NOTES = [
    "Missing data: blank cells and REDCap's missing-data codes "
    "(-666, -777, -888, -999, 666) are counted as missing. They are never averaged "
    "and never shown as a category of a variable.",
    "Percentages: the denominator is the records the field was actually asked of — "
    "its branching logic fired — and that have a usable answer. It is not always "
    "the total number of records, and for a field behind branching logic it should "
    "not be.",
    "Rounding: every statistic that is not a count is rounded to 2 decimal places.",
]

_ILLEGAL_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _openpyxl():
    return core.require("openpyxl", "openpyxl",
                        "writes Excel workbooks without needing Excel installed")


def sheet_name(name: str, taken=()) -> str:
    """An Excel-legal sheet name: 31 characters, none of []:*?/\\, and unique."""
    clean = _ILLEGAL_SHEET_CHARS.sub("-", str(name)).strip() or "Sheet"
    clean = clean[:31]
    if clean not in taken:
        return clean
    for suffix in range(2, 100):
        candidate = f"{clean[:31 - len(str(suffix)) - 1]} {suffix}"
        if candidate not in taken:
            return candidate
    return clean[:31]


def _cell_value(value):
    """What actually goes in the cell. Blanks stay blank; numbers stay numbers."""
    if value is None:
        return None
    if isinstance(value, float) and value != value:      # NaN
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    return None if text.strip() == "" else text


def _records_in(tables):
    """Find N without being told, if one of the tables is a Table 1.

    table1 always writes a `records / n` row; reading it here means the Notes
    sheet states the study size even when the caller forgot to say it.
    """
    for table in tables.values():
        columns = list(table.columns)
        if not {"variable", "statistic"} <= set(columns):
            continue
        last = "overall" if "overall" in columns else columns[-1]
        for _, row in table.iterrows():
            if str(row["variable"]) == "records" and str(row["statistic"]) == "n":
                try:
                    return int(row[last])
                except (TypeError, ValueError):
                    return None
    return None


def write_workbook(tables: dict, path, notes=None, n=None):
    """Write one workbook in the house style. Returns the path written.

    tables  {sheet name: DataFrame}, written in the order given
    path    where to write it (the folder is created if it is not there)
    notes   extra lines for the Notes sheet — what this analysis is, which
            records were excluded, anything a reader would otherwise ask
    n       how many records the analysis covers; read from a Table 1 if omitted
    """
    openpyxl = _openpyxl()
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    if not tables:
        raise ValueError("There are no tables to write — the workbook would be empty.")

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    bold = Font(bold=True)

    used = []
    for name, table in tables.items():
        title = sheet_name(name, used)
        used.append(title)
        sheet = workbook.create_sheet(title)
        headers = [str(c) for c in table.columns]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = bold
            cell.alignment = Alignment(vertical="center")
        for _, row in table.iterrows():
            sheet.append([_cell_value(v) for v in row.tolist()])

        sheet.freeze_panes = "A2"          # the header stays on screen
        widths = [len(h) for h in headers]
        for _, row in table.iterrows():
            for i, value in enumerate(row.tolist()):
                shown = "" if _cell_value(value) is None else str(_cell_value(value))
                widths[i] = max(widths[i], len(shown))
        for i, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 9), 60)

    # --- the Notes sheet, always last -------------------------------------
    sheet = workbook.create_sheet(sheet_name("Notes", used))
    sheet.append(["Notes"])
    sheet["A1"].font = Font(bold=True, size=13)
    sheet.append([])

    if n is None:
        n = _records_in(tables)
    lines = []
    if n is not None:
        lines.append(f"Records in this analysis (N): {n}")
    lines += [
        f"Produced by: {_generating_script()}",
        f"Produced on: {datetime.date.today().isoformat()}",
    ]
    lines += list(notes or [])
    lines += STANDING_NOTES
    if core.UNPARSEABLE_LOGIC:
        lines.append(
            "Conditions that could not be read, so every record was counted as having "
            "been asked the field they control — check these by hand: "
            + "; ".join(core.UNPARSEABLE_LOGIC)
        )
    for line in lines:
        sheet.append([line])
    sheet.column_dimensions["A"].width = 110
    for row in sheet.iter_rows(min_row=3):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(path)
    return path


def _generating_script() -> str:
    """The script that made this file — the provenance a reader needs to redo it."""
    argv0 = sys.argv[0] if sys.argv else ""
    if not argv0 or argv0 in ("-c", "-"):
        return "an interactive session (no script file)"
    return Path(argv0).name


def _main() -> int:
    print(__doc__.strip())
    print("\nThis file is part of the ARGO analysis library. It is not run on its own —")
    print("an analysis script imports it. See the run-analysis skill for how to start one.")
    return 0


if __name__ == "__main__":
    sys.exit(_main())
