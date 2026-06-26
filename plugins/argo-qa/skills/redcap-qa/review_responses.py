"""Diff an RA response workbook against the original highlighted worklist.

Reports cells where:
  - The original cell was yellow (RA was asked to resolve it)
  - The response now has a non-blank value

Per-cell output: site, record id, field (human label), original value, new value.

Usage:
  python3 review_responses.py <original.xlsx> <response.xlsx>
"""

from __future__ import annotations

import sys
from openpyxl import load_workbook


YELLOW_HEX = "FFC7CE"   # matches build_worklists.py


def _row_to_dict(ws, header_row=1, prereq_row=2):
    headers = [c.value for c in ws[header_row]]
    rows = []
    for r in range(prereq_row + 1, ws.max_row + 1):
        cells = [ws.cell(row=r, column=c) for c in range(1, ws.max_column + 1)]
        rows.append({"row": r, "cells": cells, "headers": headers})
    return rows, headers


def _yellow_keys(orig_ws, id_col_count=2):
    """Return {(record_id, field_label): original_value} for every yellow cell."""
    headers = [c.value for c in orig_ws[1]]
    id_field = headers[0]
    out = {}
    for r in range(3, orig_ws.max_row + 1):
        rid = str(orig_ws.cell(row=r, column=1).value or "").strip()
        if not rid:
            continue
        for c in range(id_col_count + 1, orig_ws.max_column + 1):
            cell = orig_ws.cell(row=r, column=c)
            fill = cell.fill
            color = None
            if fill and fill.fgColor:
                color = fill.fgColor.rgb
            if color and str(color).upper().endswith(YELLOW_HEX):
                out[(rid, headers[c-1])] = ("" if cell.value is None else str(cell.value))
    return id_field, headers, out


# Substrings (lowercased) that mark a column as an RA-comment / response column.
# Use substring match so variants like "RA COMMENT", "Comments", "Notes from RA",
# "RA Response" all get picked up.
RESPONSE_HEADER_TOKENS = ("response", "comment", "note")


def diff(orig_path: str, resp_path: str):
    orig_wb = load_workbook(orig_path)
    resp_wb = load_workbook(resp_path, data_only=True)
    orig_ws = orig_wb.active
    resp_ws = resp_wb.active

    id_field, orig_headers, yellow_map = _yellow_keys(orig_ws)
    resp_headers = [c.value for c in resp_ws[1]]

    # Find any RA-added RESPONSE/comment column (substring-match on header tokens)
    response_col_idx = None
    for i, h in enumerate(resp_headers, 1):
        if h and any(t in str(h).lower() for t in RESPONSE_HEADER_TOKENS):
            response_col_idx = i
            break

    # Build resp lookup by (rid, header) and a per-record RESPONSE note
    resp_lookup = {}
    response_notes = {}
    for r in range(3, resp_ws.max_row + 1):
        rid = str(resp_ws.cell(row=r, column=1).value or "").strip()
        if not rid:
            continue
        if response_col_idx:
            v = resp_ws.cell(row=r, column=response_col_idx).value
            response_notes[rid] = ("" if v is None else str(v).strip())
        for c in range(1, resp_ws.max_column + 1):
            if c == response_col_idx:
                continue
            h = resp_headers[c-1]
            v = resp_ws.cell(row=r, column=c).value
            resp_lookup[(rid, h)] = ("" if v is None else str(v).strip())

    # Group by record so the RESPONSE note + all changed cells appear together
    by_record = {}
    for (rid, field), orig_val in yellow_map.items():
        new_val = resp_lookup.get((rid, field), "")
        if new_val == "" or new_val == orig_val:
            continue
        by_record.setdefault(rid, []).append((field, orig_val, new_val))

    return by_record, response_notes, id_field, response_col_idx is not None


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(2)
    orig, resp = sys.argv[1], sys.argv[2]
    by_record, notes, id_field, had_response_col = diff(orig, resp)
    print(f"Original: {orig}")
    print(f"Response: {resp}")
    print(f"RESPONSE column present: {had_response_col}")
    print(f"{len(by_record)} records with proposed updates")
    print("=" * 100)
    for rid in sorted(by_record):
        note = notes.get(rid, "")
        print(f"\n{id_field}: {rid}")
        if note:
            print(f"  RA note: {note}")
        for field, ov, nv in by_record[rid]:
            print(f"    {field}")
            print(f"      was: {ov!r}")
            print(f"      now: {nv!r}")
    # Records with only a RESPONSE note (no cell changes) — easy to overlook
    note_only = sorted(rid for rid, n in notes.items() if n and rid not in by_record)
    if note_only:
        print("\n" + "=" * 100)
        print(f"{len(note_only)} record(s) with RA notes but no cell changes:")
        for rid in note_only:
            print(f"  {rid}: {notes[rid]}")


if __name__ == "__main__":
    main()
