"""Build per-site Excel worklists for RAs from a REDCap project.

Generalized from REDCap/Analysis/linkages/R01_linkages/build_ra_worklists.py.

For each configured workbook (a named bundle of fields), this:
  1. Pulls raw records + metadata via the REDCap API.
  2. Evaluates each field's branching logic per record to decide applicability.
  3. Flags applicable-but-blank cells (and optionally MDC sentinels
     -666/-777/-888/-999, plus 666 = N/A).
  4. Splits output by DAG (`redcap_data_access_group`) and writes one .xlsx
     per (DAG, workbook) into `<out>/with_MDC/` and `<out>/no_MDC/`.

Highlighted yellow cells = the RA needs to resolve these in REDCap. A second
header row shows each field's branching prerequisite in plain language.

Usage:
  python build_worklists.py \\
      --url $REDCAP_URL \\
      --token-env CRC_TOKEN \\
      --fields fields.yaml \\
      --out outputs/qa_worklists \\
      [--scope-ids ids.csv]            # optional: limit to these record IDs
      [--id-field record_id]           # record-id field name (default: record_id)
      [--extra-id-cols research_number,collaboration_identifier]

fields.yaml:
  workbooks:
    - name: clinical
      title: Clinical
      fields: [biopsy, biopsy_site, m_score, treatment_received, ...]
    - name: followup
      title: Follow-up
      fields: [last_followup_status, death_date1, recur1, ...]
"""

from __future__ import annotations

import argparse
import datetime as _dt
import os
import re
import sys
from io import StringIO

import pandas as pd
import requests
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# -----------------------------------------------------------------------------
# REDCap API
# -----------------------------------------------------------------------------

def redcap_post(url: str, token: str, **fields) -> str:
    payload = {"token": token, "format": "csv", "returnFormat": "json", **fields}
    r = requests.post(url, data=payload, timeout=300)
    r.raise_for_status()
    return r.text


def pull_records(url: str, token: str) -> pd.DataFrame:
    csv_text = redcap_post(url, token, content="record", type="flat",
                            rawOrLabel="raw", exportDataAccessGroups="true")
    return pd.read_csv(StringIO(csv_text), dtype=str, keep_default_na=False)


def pull_metadata(url: str, token: str) -> list[dict]:
    csv_text = redcap_post(url, token, content="metadata")
    return pd.read_csv(StringIO(csv_text), dtype=str, keep_default_na=False).to_dict("records")


# -----------------------------------------------------------------------------
# Metadata-driven branching + labels (vendored from R01_linkages/pipeline.py)
# -----------------------------------------------------------------------------

_CLAUSE_RE = re.compile(
    r"""\s*\[([a-zA-Z0-9_]+)(?:\((-?\w+)\))?\]\s*(=|<>|!=)\s*['"]([^'"]*)['"]\s*"""
)
_CLAUSE_RE_LOOSE = re.compile(
    r"\s*\[(\w+)(?:\((-?\w+)\))?\]\s*(=|<>|!=|>=|<=|>|<)\s*['\"]?([^'\"\s]+)['\"]?\s*"
)


def parse_choices(raw: str) -> dict:
    out = {}
    for part in (raw or "").split("|"):
        if "," in part:
            code, label = part.split(",", 1)
            out[code.strip()] = label.strip()
    return out


def clean_label(label: str) -> str:
    if not label:
        return label
    return re.sub(r"\s+", " ", label).strip().rstrip(" ?.:;")


def _eval_clause(clause: str, row: dict) -> bool:
    m = _CLAUSE_RE.fullmatch(clause)
    if not m:
        return False
    field, choice, op, val = m.group(1), m.group(2), m.group(3), m.group(4)
    if choice:
        suf = f"____{choice[1:]}" if choice.startswith("-") else f"___{choice}"
        col = f"{field}{suf}"
    else:
        col = field
    actual = str(row.get(col, ""))
    return actual == val if op == "=" else actual != val


def eval_branching(logic: str, row: dict) -> bool:
    """Evaluate REDCap branching logic. Fail-open on unparseable clauses."""
    if not logic or not logic.strip():
        return True
    for or_part in re.split(r"\s+OR\s+", logic, flags=re.IGNORECASE):
        ok = True
        for clause in re.split(r"\s+AND\s+", or_part, flags=re.IGNORECASE):
            if not _CLAUSE_RE.fullmatch(clause) or not _eval_clause(clause, row):
                ok = False
                break
        if ok:
            return True
    return False


def extract_branching_triggers(logic: str) -> list:
    if not logic:
        return []
    out, seen = [], set()
    for clause in re.split(r"\s+(?:AND|OR)\s+", logic, flags=re.IGNORECASE):
        m = _CLAUSE_RE.fullmatch(clause)
        if m and m.group(1) not in seen:
            seen.add(m.group(1)); out.append(m.group(1))
    return out


def augment_fields_with_triggers(fields: list, metadata: list) -> tuple[list, dict]:
    """Pull in any gate fields referenced by branching logic on `fields`."""
    meta_by = {m["field_name"]: m for m in metadata}
    prereq, augmented, seen = {}, list(fields), {f.split("___")[0] for f in fields}
    for f in fields:
        base = f.split("___")[0]
        m = meta_by.get(base)
        if not m:
            continue
        triggers = extract_branching_triggers(m.get("branching_logic", "") or "")
        if triggers:
            prereq[base] = triggers
        for t in triggers:
            if t not in seen:
                augmented.append(t); seen.add(t)
    return augmented, prereq


def format_branching_logic(logic: str, meta_by: dict) -> str:
    """Render branching logic with choice labels for the RA-facing prereq row."""
    if not logic:
        return ""
    parts = re.split(r"(\s+(?:AND|OR)\s+)", logic, flags=re.IGNORECASE)
    out = []
    for tok in parts:
        if re.fullmatch(r"\s+(?:AND|OR)\s+", tok, flags=re.IGNORECASE):
            out.append(tok.strip().upper()); continue
        m = _CLAUSE_RE_LOOSE.fullmatch(tok)
        if not m:
            out.append(tok.strip()); continue
        field, choice_code, op, val = m.group(1), m.group(2), m.group(3), m.group(4)
        choices = parse_choices(meta_by.get(field, {}).get("select_choices_or_calculations", "") or "")
        if choice_code is not None:
            label = choices.get(choice_code, choice_code)
            verb = "includes" if val == "1" else "excludes"
            out.append(f"{field} {verb} {label}")
        else:
            label = choices.get(val, val)
            out.append(f"{field} {op} {label}")
    return " ".join(out)


def labelize(df: pd.DataFrame, metadata: list, fields: list) -> tuple[pd.DataFrame, dict]:
    """Return (labeled_df, label_map: raw_field -> human header)."""
    meta_by = {m["field_name"]: m for m in metadata}
    out = df.copy()
    bases, seen = [], set()
    for f in fields:
        b = f.split("___")[0]
        if b not in seen and b in meta_by:
            bases.append(b); seen.add(b)
    label_map = {}
    for base in bases:
        m = meta_by[base]
        ftype = m.get("field_type", "")
        choices = parse_choices(m.get("select_choices_or_calculations", ""))
        label = clean_label(m.get("field_label", "")) or base

        if ftype == "checkbox":
            cols = [c for c in out.columns if c == base or c.startswith(f"{base}___")]
            if not cols:
                continue

            def _code(col, b=base):
                suf = col[len(b):]
                if suf.startswith("____"): return "-" + suf[4:]
                if suf.startswith("___"):  return suf[3:]
                return ""

            def _labels(row, cols=cols, choices=choices, _code=_code):
                return ", ".join(choices.get(_code(c), _code(c))
                                 for c in cols if str(row.get(c, "")) == "1")

            out[base] = out.apply(_labels, axis=1)
            out = out.drop(columns=[c for c in cols if c != base])
        elif ftype in ("radio", "dropdown", "yesno", "truefalse") and base in out.columns and choices:
            out[base] = out[base].map(lambda v: choices.get(str(v), v))

        label_map[base] = label
    return out, label_map


# -----------------------------------------------------------------------------
# Worklist builder
# -----------------------------------------------------------------------------

SENTINEL_CODES = {"666", "-666", "-777", "-888", "-999"}

YELLOW          = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL     = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
RESPONSE_HEADER = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
HEADER_FONT     = Font(bold=True, color="FFFFFF")
PREREQ_FONT     = Font(italic=True, size=9, color="666666")


def _option_code(col: str, base: str) -> str:
    suf = col[len(base):]
    if suf.startswith("____"): return "-" + suf[4:]
    if suf.startswith("___"):  return suf[3:]
    return ""


def missing_mask_for(field: str, raw_row: pd.Series, meta_by: dict,
                      flag_sentinels: bool = True) -> bool:
    m = meta_by.get(field)
    if not m:
        return False
    if not eval_branching(m.get("branching_logic") or "", raw_row.to_dict()):
        return False
    ftype = m.get("field_type", "")
    if ftype == "checkbox":
        ticked = [
            _option_code(k, field)
            for k in raw_row.index
            if (k == field or k.startswith(f"{field}___")) and str(raw_row.get(k, "")) == "1"
        ]
        if not ticked:
            return True
        return flag_sentinels and all(t in SENTINEL_CODES for t in ticked)
    val = str(raw_row.get(field, "")).strip()
    if val == "":
        return True
    return flag_sentinels and val in SENTINEL_CODES


def build_workbook(labeled: pd.DataFrame, raw_by_id: dict, meta_by: dict,
                    prereq_map: dict, fields: list, label_map: dict,
                    id_cols: list, title: str, flag_sentinels: bool) -> Workbook | None:
    """One workbook = wide table, applicable-but-blank cells in yellow."""
    rows_with_work, fields_with_work = [], set()
    join_id = id_cols[0]
    for _, lrow in labeled.iterrows():
        rid = str(lrow.get(join_id, "")).strip()
        rrow = raw_by_id.get(rid)
        if rrow is None:
            continue
        missing = [f for f in fields if missing_mask_for(f, rrow, meta_by, flag_sentinels)]
        if missing:
            rows_with_work.append((lrow, rrow, set(missing)))
            fields_with_work.update(missing)

    if not rows_with_work:
        return None

    # Gate context — surface fields that *gate* something flagged.
    context_set = {g for f in fields_with_work for g in prereq_map.get(f, [])
                   if g not in fields_with_work}
    display_set = fields_with_work | context_set
    display_fields = [f for f in fields if f in display_set]
    for gate in context_set:
        if gate not in display_fields:
            display_fields.insert(0, gate)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header = id_cols + [label_map.get(f, f) for f in display_fields] + ["RESPONSE"]
    prereq_row = (
        [""] * len(id_cols)
        + [
            ("only if " + format_branching_logic(
                meta_by.get(f, {}).get("branching_logic") or "", meta_by))
            if (meta_by.get(f, {}).get("branching_logic") or "").strip() else ""
            for f in display_fields
        ]
        + ["per-row notes from the RA: why blank, RESOLVED, etc."]
    )
    ws.append(header); ws.append(prereq_row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    # RESPONSE column header — distinct color so RAs notice it
    ws.cell(row=1, column=len(header)).fill = RESPONSE_HEADER
    for cell in ws[2]:
        cell.font = PREREQ_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = f"{get_column_letter(len(id_cols) + 1)}3"

    for lrow, rrow, missing in rows_with_work:
        out = [lrow.get(c, "") for c in id_cols]
        for f in display_fields:
            out.append(lrow.get(label_map.get(f, f), ""))
        out.append("")  # empty RESPONSE cell for the RA
        ws.append(out)
        row_idx = ws.max_row
        for i, f in enumerate(display_fields):
            if f in missing:
                ws.cell(row=row_idx, column=len(id_cols) + 1 + i).fill = YELLOW

    for i in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(32, len(str(header[i-1])) + 4))
    # RESPONSE column wider for free-text
    ws.column_dimensions[get_column_letter(len(header))].width = 50
    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 24
    return wb


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

def _expand_checkbox_columns(fields: list, raw_cols: list, meta_by: dict) -> list:
    out, seen = [], set()
    for f in fields:
        base = f.split("___")[0]
        m = meta_by.get(base)
        if m and m.get("field_type") == "checkbox":
            for col in raw_cols:
                if (col == base or col.startswith(f"{base}___")) and col not in seen:
                    out.append(col); seen.add(col)
        elif f in raw_cols and f not in seen:
            out.append(f); seen.add(f)
    return out


# REDCap "Download Data Dictionary" CSV (human headers) -> API metadata keys, for no-token mode.
_DD_TO_META = {
    "Variable / Field Name": "field_name", "Form Name": "form_name", "Section Header": "section_header",
    "Field Type": "field_type", "Field Label": "field_label",
    "Choices, Calculations, OR Slider Labels": "select_choices_or_calculations", "Field Note": "field_note",
    "Text Validation Type OR Show Slider Number": "text_validation_type_or_show_slider_number",
    "Text Validation Min": "text_validation_min", "Text Validation Max": "text_validation_max",
    "Identifier?": "identifier", "Branching Logic (Show field only if...)": "branching_logic",
    "Required Field?": "required_field", "Custom Alignment": "custom_alignment",
    "Question Number (surveys only)": "question_number", "Matrix Group Name": "matrix_group_name",
    "Matrix Ranking?": "matrix_ranking", "Field Annotation": "field_annotation",
}


def load_metadata_csv(path):
    """Load a local metadata CSV (no-token mode). Accepts either a REDCap API metadata export
    (already has 'field_name') or a Designer 'Download Data Dictionary' CSV (human headers)."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    if "field_name" not in df.columns:
        df = df.rename(columns=_DD_TO_META)
    return df.to_dict("records")


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--url", help="REDCap API URL (API mode)")
    ap.add_argument("--token-env", help="Env var holding the API token (API mode)")
    ap.add_argument("--records-csv", help="No-token mode: local record export CSV (use instead of --url/--token-env)")
    ap.add_argument("--metadata-csv", help="No-token mode: local Data Dictionary CSV (paired with --records-csv)")
    ap.add_argument("--fields", required=True, help="Path to fields YAML")
    ap.add_argument("--out", required=True, help="Output directory (a per-round subdir will be appended)")
    ap.add_argument("--round", dest="round_tag", default="",
                    help="Round label appended to --out so reruns don't overwrite. Defaults to today's date (YYYY-MM-DD). Pass empty string '--round=' to write to --out directly (legacy behavior).")
    ap.add_argument("--scope-ids", help="Optional CSV of record IDs to restrict to (one column, first row is header)")
    ap.add_argument("--id-field", default="record_id", help="REDCap record-id field name")
    ap.add_argument("--extra-id-cols", default="", help="Comma-separated extra ID columns to include in the worklist")
    args = ap.parse_args()
    if args.round_tag is None:
        args.round_tag = ""
    # Default round tag = today's date; explicit '' disables the subdir.
    if args.round_tag == "" and "--round=" not in " ".join(sys.argv) and "--round" not in sys.argv:
        args.round_tag = _dt.date.today().isoformat()
    if args.round_tag:
        args.out = os.path.join(args.out, args.round_tag)
        print(f"Round: {args.round_tag} → writing to {args.out}/")

    cfg = yaml.safe_load(open(args.fields))
    workbooks_cfg = cfg.get("workbooks") or []
    if not workbooks_cfg:
        sys.exit(f"No `workbooks:` in {args.fields}")

    os.makedirs(args.out, exist_ok=True)
    extra_id_cols = [c.strip() for c in args.extra_id_cols.split(",") if c.strip()]
    id_cols = [args.id_field] + extra_id_cols

    if args.records_csv:
        if not args.metadata_csv:
            sys.exit("--records-csv requires --metadata-csv (no-token mode needs the data dictionary too).")
        print(f"No-token mode: reading {args.records_csv} + {args.metadata_csv} ...")
        raw = pd.read_csv(args.records_csv, dtype=str, keep_default_na=False)
        metadata = load_metadata_csv(args.metadata_csv)
    elif args.url and args.token_env:
        token = os.environ.get(args.token_env)
        if not token:
            sys.exit(f"Env var {args.token_env} is not set")
        print(f"Pulling records + metadata from {args.url} ...")
        raw = pull_records(args.url, token)
        metadata = pull_metadata(args.url, token)
    else:
        sys.exit("Provide either --records-csv + --metadata-csv (no-token), or --url + --token-env (API).")
    meta_by = {m["field_name"]: m for m in metadata}

    if args.id_field not in raw.columns:
        sys.exit(f"--id-field {args.id_field!r} not in record columns: {list(raw.columns)[:8]}...")

    if args.scope_ids:
        scope = pd.read_csv(args.scope_ids, dtype=str)
        ids = set(scope.iloc[:, 0].astype(str).str.strip())
        raw = raw[raw[args.id_field].astype(str).str.strip().isin(ids)]
        print(f"  scoped to {len(raw)} record(s) via {args.scope_ids}")

    if "redcap_data_access_group" not in raw.columns:
        raw["redcap_data_access_group"] = "unassigned"
    raw["redcap_data_access_group"] = raw["redcap_data_access_group"].replace("", "unassigned").fillna("unassigned")

    raw_by_id = {str(r[args.id_field]).strip(): r for _, r in raw.iterrows()}
    dags = sorted(raw["redcap_data_access_group"].unique())
    print(f"DAGs: {dags}")

    for wb_cfg in workbooks_cfg:
        wb_name = wb_cfg["name"]
        title = wb_cfg.get("title", wb_name)
        fields = wb_cfg["fields"]

        augmented, prereq_map = augment_fields_with_triggers(fields, metadata)
        expanded = _expand_checkbox_columns(augmented, list(raw.columns), meta_by)
        keep_cols = [c for c in id_cols if c in raw.columns] + ["redcap_data_access_group"] + \
                    [c for c in expanded if c not in id_cols]
        sub_raw = raw[keep_cols].copy()
        labeled, label_map = labelize(sub_raw, metadata, augmented)
        # rename labeled headers to human labels
        labeled = labeled.rename(columns=label_map)

        for dag in dags:
            sub = labeled[labeled["redcap_data_access_group"] == dag]
            if sub.empty:
                continue
            for flag_sentinels, subfolder in [(True, "with_MDC"), (False, "no_MDC")]:
                wb = build_workbook(sub, raw_by_id, meta_by, prereq_map,
                                    fields, label_map, id_cols, title, flag_sentinels)
                folder = os.path.join(args.out, subfolder)
                os.makedirs(folder, exist_ok=True)
                tag = f"{wb_name}_{dag}"
                if wb is None:
                    print(f"  skip  {subfolder}/{tag}: nothing to review")
                    continue
                path = os.path.join(folder, f"{tag}.xlsx")
                wb.save(path)
                ws = wb.active
                print(f"  wrote {path}  ({ws.max_row - 2} patients × {ws.max_column - len(id_cols)} fields)")

    print("Done.")


if __name__ == "__main__":
    main()
