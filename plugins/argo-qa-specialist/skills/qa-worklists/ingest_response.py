"""Translate a freetext-label RA response workbook into a coded REDCap push CSV.

RAs often fill worklists with human-readable LABELS ("Yes", "M1", "Died of
disease", "Surgery, Chemotherapy", a CEA number, a date) rather than REDCap
codes. This script maps each filled cell back to the code(s) REDCap expects,
using the project metadata, and emits a push CSV ready for verify_push.py.

What it handles:
  - radio / dropdown / yesno : label  -> code        (case-insensitive)
  - checkbox                 : "A, B" -> base___codeA=1, base___codeB=1
  - date fields              : any parseable date -> YYYY-MM-DD
  - text / notes             : passthrough (numbers, places, CEA values…)
  - "case note not found" &  : -> -888 (Missing in case notes), or the
    similar missingness phrases   date-format 08-08-8888 on date fields

Anything it can't map confidently is written to the --report markdown instead
of being guessed — so ambiguous cells become RA questions, never silent data.

With --original <worklist.xlsx>, only cells the RA actually changed from the
worklist are translated (context cells the RA left untouched are skipped).

Usage:
  python3 ingest_response.py --url $REDCAP_URL --token-env CRC_TOKEN \\
      --response "RA_response/clinical_luth.xlsx" \\
      --out push_drafts/<round>/luth_clinical.csv \\
      [--original outputs/<round>/with_MDC/clinical_luth.xlsx] \\
      [--report push_drafts/<round>/luth_clinical_unmapped.md] \\
      [--id-field research_number]
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import re
import sys
from io import StringIO

import requests
import yaml
from openpyxl import load_workbook

# Reuse the exact label-cleaning the worklist builder uses, so headers line up.
from build_worklists import clean_label, parse_choices
# ...and the exact fill the builder paints. A copy of the hex here that drifted from the
# builder's would mean every flagged cell stops being recognised, silently. See qa_colours.py.
from qa_colours import YELLOW_HEX

# Normalized response substrings that mean "RA looked, it's missing from the chart".
# Kept deliberately loose to absorb RA typos ("case not cannot be found").
MDC_MISSING_PATTERNS = (
    "not found", "cannot be found", "could not be found", "can not be found",
    "no case note", "case note missing", "missing case note",
    "missing in case note", "case note not available", "case note unavailable",
    "not available in case note", "no case file", "case file not",
)
MDC_MISSING_CODE = "-888"
MDC_MISSING_DATE = "08-08-8888"   # date-format equivalent of -888

RESPONSE_HEADER_TOKENS = ("response", "comment", "note")


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def load_metadata(url: str, token: str) -> dict:
    r = requests.post(url, data={"token": token, "content": "metadata",
                                  "format": "csv", "returnFormat": "json"}, timeout=120)
    r.raise_for_status()
    return {m["field_name"]: m for m in csv.DictReader(StringIO(r.text))}


def build_label_maps(meta_by: dict):
    """label2fields: cleaned field_label -> [field_name, ...] (in data-dictionary order).
       value2code[field]: normalized choice label -> code.

    A list, not a single name: labels are NOT unique in REDCap — only field names are. A live
    160-field dictionary had 44 labels shared by more than one field. Keeping just the first
    meant every cell under a shared heading was pushed to whichever field happened to come
    first in the dictionary, silently and with no way to notice.
    """
    label2fields, value2code = {}, {}
    for fname, m in meta_by.items():
        lbl = clean_label(m.get("field_label", ""))
        if lbl:
            label2fields.setdefault(_norm(lbl), []).append(fname)
        choices = parse_choices(m.get("select_choices_or_calculations", ""))
        if choices:
            value2code[fname] = {_norm(lbl): code for code, lbl in choices.items()}
    return label2fields, value2code


# build_worklists.display_headers writes a repeated label as "Label (field_name)".
_DISAMBIGUATED_HEADER = re.compile(r"^(?P<label>.*)\s\((?P<field>[A-Za-z0-9_]+)(?:\s#\d+)?\)$")


def header_to_field(header, label2fields: dict, meta_by: dict) -> "tuple[str | None, str]":
    """Which REDCap field a worklist column belongs to: (field_name, reason_if_none).

    Tried in order of how certain each answer is:
      1. the column heading IS a field name  — a worklist written without labels;
      2. the heading is "Label (field_name)" — how the builder disambiguates a repeated label;
      3. the heading is a label used by exactly one field.
    A label shared by several fields resolves to none of them. We do not pick one: writing a
    value to the wrong REDCap field is worse than not writing it, and the unmapped report says
    which columns need a human.
    """
    raw = str(header or "").strip()
    if not raw:
        return None, "blank column heading"
    if raw in meta_by:
        return raw, ""
    m = _DISAMBIGUATED_HEADER.match(raw)
    if m and m.group("field") in meta_by:
        return m.group("field"), ""
    candidates = label2fields.get(_norm(raw), [])
    if len(candidates) == 1:
        return candidates[0], ""
    if len(candidates) > 1:
        return None, ("this label is shared by " + ", ".join(candidates) +
                      " — rebuild the worklist so the heading names the field")
    return None, "no REDCap field has this label"


def _is_missing_phrase(v: str) -> bool:
    n = _norm(v)
    return any(p in n for p in MDC_MISSING_PATTERNS)


def _parse_date(v) -> str | None:
    if isinstance(v, dt.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, dt.date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d",
                "%d.%m.%Y", "%d %b %Y", "%d %B %Y", "%b %d %Y", "%B %d %Y"):
        try:
            return dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _cb_col(base: str, code: str) -> str:
    code = str(code)
    return f"{base}____{code[1:]}" if code.startswith("-") else f"{base}___{code}"


def _resolve_alias(field: str, part_norm: str, aliases: dict, value2code: dict):
    """Map an RA phrase via the alias table to a code. Alias targets may be a
    canonical choice label (resolved through value2code) or a literal code."""
    target = aliases.get(field, {}).get(part_norm)
    if target is None:
        return None
    # alias value could be a label -> resolve to code, or already be a code
    code = value2code.get(field, {}).get(_norm(target))
    if code is not None:
        return code
    return str(target)   # treat as a literal code


def translate_cell(field: str, value, meta_by: dict, value2code: dict, aliases: dict,
                    issues: list, rid: str, label: str) -> dict:
    """Return {column: value} updates for one (field, value), or {} if unmappable.

    Unmappable / ambiguous values are appended to `issues` and skipped."""
    m = meta_by.get(field, {})
    ftype = m.get("field_type", "")
    is_date = (m.get("text_validation_type_or_show_slider_number", "") or "").startswith("date")

    # Missingness phrase first — interpretation depends on field type.
    if _is_missing_phrase(value):
        if ftype == "checkbox":
            return {_cb_col(field, MDC_MISSING_CODE): "1"}
        if is_date:
            return {field: MDC_MISSING_DATE}
        return {field: MDC_MISSING_CODE}

    if ftype == "checkbox":
        out = {}
        for part in str(value).split(","):
            pn = _norm(part)
            if not pn:
                continue
            code = value2code.get(field, {}).get(pn) or _resolve_alias(field, pn, aliases, value2code)
            if code is None:
                issues.append((rid, label, str(value), f"checkbox option {part.strip()!r} not a choice"))
                continue
            out[_cb_col(field, code)] = "1"
        return out

    if ftype in ("radio", "dropdown", "yesno", "truefalse"):
        code = value2code.get(field, {}).get(_norm(value)) or _resolve_alias(field, _norm(value), aliases, value2code)
        if code is None:
            issues.append((rid, label, str(value), "value not in choice list"))
            return {}
        return {field: code}

    # text / notes / etc.
    if is_date:
        d = _parse_date(value)
        if d is None:
            issues.append((rid, label, str(value), "unparseable date"))
            return {}
        return {field: d}

    # plain text/number — passthrough, but normalize openpyxl ints/floats
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return {field: str(value).strip()}


def _changed_cells(original_path: str, id_field_out: str):
    """{(rid, header): original_value} for yellow cells, plus all headers — used to
    restrict translation to cells the RA actually changed."""
    if not original_path:
        return None
    ws = load_workbook(original_path).active
    headers = [c.value for c in ws[1]]
    orig = {}
    for r in range(3, ws.max_row + 1):
        rid = str(ws.cell(row=r, column=1).value or "").strip()
        if not rid:
            continue
        for c in range(3, ws.max_column + 1):
            orig[(rid, headers[c-1])] = ("" if ws.cell(row=r, column=c).value is None
                                          else str(ws.cell(row=r, column=c).value).strip())
    return orig


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--url", required=True)
    ap.add_argument("--token-env", required=True)
    ap.add_argument("--response", required=True, help="RA response workbook (.xlsx)")
    ap.add_argument("--out", required=True, help="Output push CSV")
    ap.add_argument("--original", help="Original worklist — restrict to cells the RA changed")
    ap.add_argument("--aliases", help="YAML of {field: {ra_phrase: canonical_label_or_code}} for common RA wordings that don't exactly match REDCap choices")
    ap.add_argument("--report", help="Markdown file for unmapped/ambiguous cells (default: <out>_unmapped.md)")
    ap.add_argument("--id-field", default="research_number", help="REDCap record-id field name (output column)")
    args = ap.parse_args()

    tok = os.environ.get(args.token_env)
    if not tok:
        sys.exit(
        f"No access key for {args.token_env} is set up on this computer, so I can't reach REDCap.\n"
        "\n"
        "An access key (REDCap calls it an API token) is a long password that lets a tool read\n"
        "or update one specific REDCap project on your behalf. Your REDCap administrator\n"
        "creates it for you — it isn't something you can generate yourself.\n"
        "\n"
        "If you already have one, add it to the file ~/.argo/.env, then load it into this\n"
        "terminal window and try again:\n"
        "\n"
        "    set -a; source ~/.argo/.env; set +a"
    )
    report_path = args.report or (os.path.splitext(args.out)[0] + "_unmapped.md")

    aliases = {}
    if args.aliases:
        raw = yaml.safe_load(open(args.aliases)) or {}
        # normalize the phrase keys for matching
        aliases = {fld: {_norm(k): v for k, v in (m or {}).items()} for fld, m in raw.items()}

    meta_by = load_metadata(args.url, tok)
    label2fields, value2code = build_label_maps(meta_by)

    ws = load_workbook(args.response, data_only=True).active
    headers = [c.value for c in ws[1]]

    # Identify id column and any RESPONSE/comment column to ignore.
    response_cols = {i for i, h in enumerate(headers, 1)
                     if h and any(t in str(h).lower() for t in RESPONSE_HEADER_TOKENS)}
    # second id column (collaboration_identifier / r01_record_id) is column 2 by convention
    id_cols = {1, 2}

    orig = _changed_cells(args.original, args.id_field)

    out_rows = []
    issues: list = []
    unknown_cols = {}
    for r in range(3, ws.max_row + 1):
        rid = str(ws.cell(row=r, column=1).value or "").strip()
        if not rid:
            continue
        row_update = {args.id_field: rid}
        for c in range(1, ws.max_column + 1):
            if c in id_cols or c in response_cols:
                continue
            header = headers[c-1]
            value = ws.cell(row=r, column=c).value
            if value in (None, ""):
                continue
            # restrict to changed cells if an original was given
            if orig is not None:
                prev = orig.get((rid, header))
                if prev is not None and str(value).strip() == prev:
                    continue
            field, why = header_to_field(header, label2fields, meta_by)
            if not field:
                unknown_cols[str(header)] = why
                continue
            updates = translate_cell(field, value, meta_by, value2code, aliases, issues, rid, header)
            row_update.update(updates)
        if len(row_update) > 1:
            out_rows.append(row_update)

    # Write push CSV (union of all columns, id first)
    cols = [args.id_field]
    for row in out_rows:
        for k in row:
            if k not in cols:
                cols.append(k)
    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    with open(args.out, "w") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for row in out_rows:
            w.writerow({c: row.get(c, "") for c in cols})
    print(f"Wrote {args.out}: {len(out_rows)} record(s), {len(cols)-1} field-columns")

    # Write the unmapped/ambiguous report
    if issues or unknown_cols:
        lines = [f"# Unmapped cells — {os.path.basename(args.response)}", ""]
        if unknown_cols:
            lines += ["## Columns that couldn't be matched to a REDCap field", ""]
            lines += [f"- {c!r} — {unknown_cols[c]}" for c in sorted(unknown_cols)] + [""]
        if issues:
            lines += ["## Values that couldn't be mapped to a code", "",
                       "| record | field | RA wrote | why |", "|---|---|---|---|"]
            for rid, label, raw, why in issues:
                lines.append(f"| {rid} | {label} | `{raw}` | {why} |")
        open(report_path, "w").write("\n".join(lines) + "\n")
        print(f"  ⚠️  {len(issues)} unmapped value(s), {len(unknown_cols)} unknown column(s) → {report_path}")
    else:
        print("  no unmapped cells — clean translation")


if __name__ == "__main__":
    main()
