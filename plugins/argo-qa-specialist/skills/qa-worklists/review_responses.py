"""Diff an RA response workbook against the original highlighted worklist.

Reports, grouped by record:
  - ANSWERS: a cell the worklist flagged (yellow "this applies and is blank", or amber
    "we couldn't read this field's condition — please check") that now holds a different,
    non-blank value. Amber answers are reported and tagged as such: an answer in an amber
    cell is still an answer.
  - The RA's per-row RESPONSE / comment note.
  - OUT-OF-SCOPE EDITS: any other cell the RA changed — a gate-context column, an ID
    column, a field that was never flagged. These are reported separately because they are
    a different kind of event: nobody asked for them, and they may be corrections, may be
    accidents, and are never safe to treat as an answer to a question we asked.

A worklist built by an older version of ARGO highlighted its gaps in a rose fill rather than
yellow. Those cells are read exactly like yellow ones — the RA answered the same question — and
a workbook mostly painted that way gets one line saying so.

Per-cell output: record id, field (human label), original value, new value.

Usage:
  python3 review_responses.py <original.xlsx> <response.xlsx>
"""

from __future__ import annotations

import re
import sys
import zipfile
from pathlib import Path
from typing import NamedTuple

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# Same-folder imports, always: this skill carries its own copy of everything it needs.
sys.path.insert(0, str(Path(__file__).resolve().parent))


def open_worklist(path: str, role: str, **kwargs):
    """Open an Excel worklist, explaining clearly what to do when it can't be opened.

    `role` is how to refer to the file in a message, e.g. "original worklist".
    """
    p = Path(path).expanduser()
    if not p.exists():
        raise SystemExit(
            f"I couldn't find the {role}:\n"
            f"    {p}\n"
            "\n"
            "Check the file name and the folder are right. If the name has spaces in it, put\n"
            'quotation marks around it, like "Ife site worklist.xlsx".'
        )
    if p.is_dir():
        raise SystemExit(
            f"The {role} path points at a folder, not a file:\n"
            f"    {p}\n"
            "\n"
            "Give me the .xlsx file inside that folder instead."
        )
    if p.suffix.lower() != ".xlsx":
        raise SystemExit(
            f"The {role} needs to be an Excel .xlsx file, but this one is {p.suffix or 'not'}:\n"
            f"    {p}\n"
            "\n"
            "If an RA sent it back in a different format, open it in Excel and use\n"
            'File → Save As → Excel Workbook (.xlsx), then try again.'
        )
    try:
        return load_workbook(p, **kwargs)
    except zipfile.BadZipFile:
        raise SystemExit(
            f"The {role} can't be opened — the file looks damaged or incomplete:\n"
            f"    {p}\n"
            "\n"
            "This usually means it didn't download or copy fully, or it was renamed to .xlsx\n"
            "without actually being an Excel file. Try opening it in Excel: if Excel can't open\n"
            "it either, ask whoever sent it for a fresh copy."
        )
    except Exception as e:
        raise SystemExit(
            f"Something went wrong opening the {role}:\n"
            f"    {p}\n"
            f"\n{type(e).__name__}: {e}\n"
            "\n"
            "Try opening the file in Excel to check it's intact."
        )


def _row_to_dict(ws, header_row=1, prereq_row=2):
    headers = [c.value for c in ws[header_row]]
    rows = []
    for r in range(prereq_row + 1, ws.max_row + 1):
        cells = [ws.cell(row=r, column=c) for c in range(1, ws.max_column + 1)]
        rows.append({"row": r, "cells": cells, "headers": headers})
    return rows, headers


# The fills the builder paints, from the one place they're defined. Retyping them here is
# how a returned workbook stops being recognised at all — every RA answer in it silently
# discarded — so this imports rather than copies. See qa_colours.py.
from qa_colours import AMBER_HEX, LEGACY_FLAG_HEXES, YELLOW_HEX  # noqa: E402

# Fill colour -> what the worklist was asking the RA to do.
#
# LEGACY_FLAG_HEXES read as "yellow" — the same question, asked in the colour an older release
# painted. A site that received a worklist before the colour changed sends it back in the old
# rose months later; the RA did the work either way. Reading only the current yellow made one
# live round report 5 of 36 answers and say nothing about the other 31.
FLAG_KINDS = {YELLOW_HEX: "yellow", AMBER_HEX: "amber"}
FLAG_KINDS.update({h: "yellow" for h in LEGACY_FLAG_HEXES})
LEGACY_HEXES = frozenset(LEGACY_FLAG_HEXES)

# Warn about the old colour once, when it's how the workbook was mostly painted — not on a
# stray cell somebody recoloured by hand. Strictly more than half.
LEGACY_WARNING_SHARE = 0.5

# Substrings (lowercased) that mark a column as an RA-comment / response column.
# Use substring match so variants like "RA COMMENT", "Comments", "Notes from RA",
# "RA Response" all get picked up.
RESPONSE_HEADER_TOKENS = ("response", "comment", "note")


class Answer(NamedTuple):
    """One flagged cell the RA answered."""
    field: str
    was: str
    now: str
    kind: str          # "yellow" (confirmed gap) or "amber" (condition unreadable)


class OutOfScopeEdit(NamedTuple):
    """One cell the RA changed that the worklist never asked about."""
    record: str
    field: str
    was: str
    now: str
    is_id_column: bool


class Audit(NamedTuple):
    by_record: dict            # rid -> [Answer, ...]
    notes: dict                # rid -> RA's RESPONSE note
    id_field: str              # header of column 1
    has_response_col: bool
    out_of_scope: list         # [OutOfScopeEdit, ...]
    flags_total: int = 0       # cells the ORIGINAL worklist highlighted, any colour
    legacy_flags: int = 0      # of those, ones painted in a retired colour


def _flag_hex(cell) -> str:
    """The known flag colour this cell is painted, or "" if it isn't painted one."""
    fill = cell.fill
    if not fill or not fill.fgColor:
        return ""
    color = str(fill.fgColor.rgb or "").upper()
    for hexv in FLAG_KINDS:
        if color.endswith(hexv):
            return hexv
    return ""


def _fill_kind(cell) -> str:
    """"yellow" / "amber" / "" for a cell, by its fill colour.

    A retired fill (see qa_colours.LEGACY_FLAG_HEXES) reads as "yellow": it asked the RA the
    same question, in the colour the builder used at the time.
    """
    return FLAG_KINDS.get(_flag_hex(cell), "")


def legacy_flag_note(flags_total: int, legacy_flags: int) -> str:
    """One line, or "" — said when a workbook's highlighting is mostly a retired colour."""
    if not flags_total or legacy_flags <= flags_total * LEGACY_WARNING_SHARE:
        return ""
    return (f"Note: {legacy_flags} of the {flags_total} highlighted cells are the old rose "
            "fill, not yellow — these worklists were built by an older version of ARGO, so "
            "I'm reading the old colour as a flag.")


def _cell_text(value) -> str:
    return "" if value is None else str(value).strip()


def id_column_count(ws) -> int:
    """How many leading ID columns this worklist has — read off the workbook, not guessed.

    `build_worklists.build_workbook` lays every worklist out the same way:

        header row 1 : <id-field> [<extra id cols>...]  <one column per field>  RESPONSE
        row 2        : blank for the ID block, "only if ..." for gated fields
        freeze_panes : <letter of len(id_cols) + 1> + "3"

    That frozen split is the workbook's own record of where the ID block ends and the data
    begins, so we read it. This used to be hardcoded to 2, and the scan for flagged cells
    started at column 3 — so on a single-ID workbook (the builder's default: `--id-field`
    with no `--extra-id-cols`) the entire first data column was invisible and every RA answer
    in it was silently discarded.

    Verified before it is trusted: if any cell to the left of the split is highlighted, the
    split is wrong for this file and we fall back to 1. Over-counting drops answers;
    under-counting cannot, because the builder never highlights an ID cell.
    """
    fallback = 1
    frozen = getattr(ws, "freeze_panes", None)
    if not frozen:
        return fallback
    m = re.fullmatch(r"([A-Za-z]+)(\d+)", str(frozen).replace("$", ""))
    if not m:
        return fallback
    try:
        n = column_index_from_string(m.group(1).upper()) - 1
    except ValueError:
        return fallback
    if not 1 <= n < ws.max_column:
        return fallback
    for r in range(3, ws.max_row + 1):
        for c in range(2, n + 1):
            if _fill_kind(ws.cell(row=r, column=c)):
                return fallback
    return n


def flagged_cells(orig_ws) -> tuple:
    """(id_field, headers, {(record_id, header): (original_value, kind)}, legacy_count).

    A flagged cell is one the worklist highlighted: yellow (applies and is blank) or amber
    (we could not read its condition — please check). Both are questions we asked the RA, so
    both count as answered when a value comes back. A cell painted a retired flag colour is
    yellow too, and `legacy_count` says how many of them there were.
    """
    headers = [c.value for c in orig_ws[1]]
    id_field = headers[0]
    first_data_col = id_column_count(orig_ws) + 1
    out, legacy = {}, 0
    for r in range(3, orig_ws.max_row + 1):
        rid = _cell_text(orig_ws.cell(row=r, column=1).value)
        if not rid:
            continue
        for c in range(first_data_col, orig_ws.max_column + 1):
            cell = orig_ws.cell(row=r, column=c)
            hexv = _flag_hex(cell)
            if not hexv:
                continue
            out[(rid, headers[c - 1])] = (_cell_text(cell.value), FLAG_KINDS[hexv])
            if hexv in LEGACY_HEXES:
                legacy += 1
    return id_field, headers, out, legacy


def _response_column(headers) -> "int | None":
    for i, h in enumerate(headers, 1):
        if h and any(t in str(h).lower() for t in RESPONSE_HEADER_TOKENS):
            return i
    return None


def diff(orig_path: str, resp_path: str) -> Audit:
    orig_wb = open_worklist(orig_path, "original worklist")
    resp_wb = open_worklist(resp_path, "worklist the RA sent back", data_only=True)
    orig_ws = orig_wb.active
    resp_ws = resp_wb.active

    id_field, orig_headers, flagged, legacy_flags = flagged_cells(orig_ws)
    id_cols = id_column_count(orig_ws)
    resp_headers = [c.value for c in resp_ws[1]]

    # Find any RA-added RESPONSE/comment column (substring-match on header tokens)
    response_col_idx = _response_column(resp_headers)

    # Every value in the ORIGINAL, by (record, header) — the baseline an out-of-scope edit
    # is measured against.
    orig_lookup = {}
    for r in range(3, orig_ws.max_row + 1):
        rid = _cell_text(orig_ws.cell(row=r, column=1).value)
        if not rid:
            continue
        for c in range(1, orig_ws.max_column + 1):
            orig_lookup[(rid, orig_headers[c - 1])] = _cell_text(orig_ws.cell(row=r, column=c).value)

    # Build resp lookup by (rid, header) and a per-record RESPONSE note
    resp_lookup = {}
    response_notes = {}
    for r in range(3, resp_ws.max_row + 1):
        rid = _cell_text(resp_ws.cell(row=r, column=1).value)
        if not rid:
            continue
        if response_col_idx:
            response_notes[rid] = _cell_text(resp_ws.cell(row=r, column=response_col_idx).value)
        for c in range(1, resp_ws.max_column + 1):
            if c == response_col_idx:
                continue
            h = resp_headers[c - 1]
            resp_lookup[(rid, h)] = _cell_text(resp_ws.cell(row=r, column=c).value)

    # Group by record so the RESPONSE note + all answered cells appear together
    by_record = {}
    for (rid, field), (orig_val, kind) in flagged.items():
        new_val = resp_lookup.get((rid, field), "")
        if new_val == "" or new_val == orig_val:
            continue
        by_record.setdefault(rid, []).append(Answer(field, orig_val, new_val, kind))
    for rid in by_record:
        by_record[rid].sort()

    # Out-of-scope edits: any OTHER cell that changed. SKILL.md promises the audit shows every
    # cell the RA changed; without this it showed only the ones we had asked about, so an RA
    # who "corrected" a gate-context column — flipping Sex, or a status that other fields
    # branch on — passed the audit in complete silence.
    out_of_scope = []
    for (rid, header), new_val in sorted(resp_lookup.items(), key=lambda kv: (kv[0][0], str(kv[0][1]))):
        if (rid, header) in flagged:
            continue
        if (rid, header) not in orig_lookup:
            continue        # a column the RA added, or a record not in the original
        old_val = orig_lookup[(rid, header)]
        if new_val == old_val:
            continue
        col = orig_headers.index(header) + 1 if header in orig_headers else 0
        out_of_scope.append(OutOfScopeEdit(rid, header, old_val, new_val,
                                           is_id_column=1 <= col <= id_cols))

    return Audit(by_record, response_notes, id_field, response_col_idx is not None, out_of_scope,
                 flags_total=len(flagged), legacy_flags=legacy_flags)


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        print(
            "This compares two Excel worklists: the one you originally sent to a site, and the\n"
            "one the RA filled in and sent back. It shows you every cell they changed.\n"
            "\n"
            "Give it both file names, original first:\n"
            "\n"
            "    python3 review_responses.py original.xlsx returned.xlsx"
        )
        sys.exit(2)
    orig, resp = sys.argv[1], sys.argv[2]
    audit = diff(orig, resp)
    by_record, notes, id_field = audit.by_record, audit.notes, audit.id_field
    amber_total = sum(1 for cells in by_record.values() for a in cells if a.kind == "amber")
    print(f"Original: {orig}")
    print(f"Response: {resp}")
    print(f"RESPONSE column present: {audit.has_response_col}")
    legacy = legacy_flag_note(audit.flags_total, audit.legacy_flags)
    if legacy:
        print(legacy)
    print(f"{len(by_record)} records with proposed updates")
    if amber_total:
        print(f"{amber_total} of the answers are in amber cells "
              "(we could not read the field's condition — check the field really applies)")
    print("=" * 100)
    for rid in sorted(by_record):
        note = notes.get(rid, "")
        print(f"\n{id_field}: {rid}")
        if note:
            print(f"  RA note: {note}")
        for ans in by_record[rid]:
            tag = "   [AMBER — we could not read this field's condition; confirm it applies]" \
                if ans.kind == "amber" else ""
            print(f"    {ans.field}{tag}")
            print(f"      was: {ans.was!r}")
            print(f"      now: {ans.now!r}")
    # Records with only a RESPONSE note (no cell changes) — easy to overlook
    note_only = sorted(rid for rid, n in notes.items() if n and rid not in by_record)
    if note_only:
        print("\n" + "=" * 100)
        print(f"{len(note_only)} record(s) with RA notes but no cell changes:")
        for rid in note_only:
            print(f"  {rid}: {notes[rid]}")
    # Cells nobody asked about. Never treat these as answers — read them, then decide.
    if audit.out_of_scope:
        print("\n" + "=" * 100)
        print(f"{len(audit.out_of_scope)} cell(s) changed that were NOT on the worklist:")
        for e in audit.out_of_scope:
            marker = "  (ID COLUMN)" if e.is_id_column else ""
            print(f"  {e.record}  {e.field}{marker}")
            print(f"      was: {e.was!r}")
            print(f"      now: {e.now!r}")
        print("\nThese were not questions we asked. Check each one against REDCap before acting")
        print("on it — an edit here can change which other fields apply.")


if __name__ == "__main__":
    main()
