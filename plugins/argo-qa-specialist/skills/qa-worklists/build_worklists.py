"""Build per-site Excel worklists for RAs from a REDCap project.

For each configured workbook (a named bundle of fields), this:
  1. Pulls raw records + metadata via the REDCap API.
  2. Evaluates each field's branching logic per record to decide applicability.
     Conditions it cannot read are surfaced and marked uncertain, never dropped.
  3. Flags applicable-but-blank cells (and optionally MDC sentinels
     -666/-777/-888/-999, plus 666 = N/A).
  4. Splits output by DAG (`redcap_data_access_group`) and writes one .xlsx
     per (DAG, workbook) into `<out>/with_MDC/` and `<out>/no_MDC/`.

Highlighted yellow cells = the RA needs to resolve these in REDCap. A second
header row shows each field's branching prerequisite in plain language.

Usage (it finds your ARGO settings file by itself — there is nothing to load first):
  python build_worklists.py \\
      --token-env CRC_TOKEN \\
      --fields fields.yaml \\
      --out qa-specialist/<study>/worklists \\
      [--scope-ids ids.csv]            # optional: limit to these record IDs
      [--id-field record_id]           # record-id field name (default: record_id)
      [--extra-id-cols research_number,collaboration_identifier]

fields.yaml:
  workbooks:
    - name: clinical
      title: Clinical
      fields: [biopsy, biopsy_site, m_score, treatment_received, ...]
    - name: followup
      title: Follow-up
      fields: [last_followup_status, death_date1, recur1, ...]
"""

from __future__ import annotations

import argparse
import datetime as _dt
import os
import re
import sys
from io import StringIO
from pathlib import Path

import pandas as pd
import requests
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Same-folder imports, always: this skill carries its own copy of everything it needs.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from qa_colours import AMBER_HEX, YELLOW_HEX  # noqa: E402

# The shared ARGO scripts are vendored into this skill's own scripts/ folder by release.py,
# so imports never depend on where — or whether — other plugins are installed. The parents walk
# is only for running from a source checkout before the first sync.
_here = Path(__file__).resolve().parent
for _cand in (_here / "scripts",
              *(p / "plugins/argo-core/skills/redcap-api/scripts" for p in _here.parents)):
    if (_cand / "argo_redcap_client.py").exists():
        sys.path.insert(0, str(_cand))
        break
from argo_redcap_client import load_env_file  # noqa: E402


# -----------------------------------------------------------------------------
# REDCap API
# -----------------------------------------------------------------------------

def redcap_post(url: str, token: str, **fields) -> str:
    payload = {"token": token, "format": "csv", "returnFormat": "json", **fields}
    r = requests.post(url, data=payload, timeout=300)
    r.raise_for_status()
    return r.text


def pull_records(url: str, token: str) -> pd.DataFrame:
    csv_text = redcap_post(url, token, content="record", type="flat",
                            rawOrLabel="raw", exportDataAccessGroups="true")
    return pd.read_csv(StringIO(csv_text), dtype=str, keep_default_na=False)


def pull_metadata(url: str, token: str) -> list[dict]:
    csv_text = redcap_post(url, token, content="metadata")
    return pd.read_csv(StringIO(csv_text), dtype=str, keep_default_na=False).to_dict("records")


# -----------------------------------------------------------------------------
# Metadata-driven branching + labels
# -----------------------------------------------------------------------------

# One clause pattern, used everywhere. It must accept everything REDCap's Designer actually
# emits, which is broader than it first appears:
#   [field] = 'value'     quoted, the form most docs show
#   [field] = 1           UNQUOTED — what REDCap writes for numeric codes, and by far the most
#                         common form in practice. A stricter pattern that required quotes was
#                         silently dropping 28% of branching fields on one live cohort and 70%
#                         on another, because an unparseable clause was treated as "not applicable".
#   [field(2)] = 1        a single checkbox option
#   [age] >= 18           numeric comparison (documented in dd-column-spec.md)
_CLAUSE_RE = re.compile(
    r"""\s*\[([a-zA-Z0-9_]+)(?:\((-?\w+)\))?\]\s*"""
    r"""(<=|>=|<>|!=|=|<|>)\s*"""
    r"""(?:'([^']*)'|"([^"]*)"|([^\s'"]+))\s*"""
)

# Logic strings we couldn't understand this run. Reported once at the end rather than per row,
# so the operator finds out the parser needs extending instead of silently losing coverage.
UNPARSEABLE_LOGIC: set = set()


def _clause_parts(clause: str):
    """(field, choice_code, operator, value) for a clause, or None if it can't be parsed."""
    m = _CLAUSE_RE.fullmatch(clause)
    if not m:
        return None
    field, choice, op = m.group(1), m.group(2), m.group(3)
    # Exactly one of the three value groups matches, depending on the quoting style.
    value = next(g for g in (m.group(4), m.group(5), m.group(6)) if g is not None)
    return field, choice, op, value


def parse_choices(raw: str) -> dict:
    out = {}
    for part in (raw or "").split("|"):
        if "," in part:
            code, label = part.split(",", 1)
            out[code.strip()] = label.strip()
    return out


def clean_label(label: str) -> str:
    if not label:
        return label
    return re.sub(r"\s+", " ", label).strip().rstrip(" ?.:;")


def _eval_clause(clause: str, row: dict) -> "bool | None":
    """True/False if the clause could be evaluated; None if it couldn't be parsed."""
    parts = _clause_parts(clause)
    if parts is None:
        return None
    field, choice, op, val = parts
    if choice:
        suf = f"____{choice[1:]}" if choice.startswith("-") else f"___{choice}"
        col = f"{field}{suf}"
    else:
        col = field
    actual = str(row.get(col, "")).strip()

    if op in ("=", "<>", "!="):
        equal = actual == str(val).strip()
        return equal if op == "=" else not equal

    # Numeric comparison.
    # A BLANK value fails the comparison, and that's a definite answer, not an unknown one:
    # REDCap evaluates `[x] >= 1` with x empty as false and hides the field, so we match it.
    # Treating blanks as "uncertain" instead would flood a worklist — on the Study Tracker it
    # would have marked a quarter of all cells "please check" for no reason.
    if actual == "":
        return False
    try:
        left, right = float(actual), float(val)
    except (TypeError, ValueError):
        return None      # a genuinely non-numeric value on either side — we can't say
    return {"<": left < right, ">": left > right,
            "<=": left <= right, ">=": left >= right}[op]


def evaluate_branching(logic: str, row: dict) -> "tuple[bool, bool]":
    """Decide whether a field applies to this record.

    Returns (applies, certain).

    `certain` is False when some part of the logic couldn't be understood. In that case we say the
    field DOES apply — never silently drop a field just because we couldn't read its condition,
    which is precisely the failure this replaced — but the caller marks it as uncertain rather
    than asserting it as a confirmed gap, and the logic string is reported at the end of the run.
    """
    if not logic or not logic.strip():
        return True, True

    any_unparseable = False
    for or_part in re.split(r"\s+OR\s+", logic, flags=re.IGNORECASE):
        branch = True
        branch_unparseable = False
        for clause in re.split(r"\s+AND\s+", or_part, flags=re.IGNORECASE):
            result = _eval_clause(clause, row)
            if result is None:
                branch_unparseable = any_unparseable = True
                continue          # unknown — don't let it decide the branch either way
            if not result:
                branch = False
                break
        if branch and not branch_unparseable:
            return True, True     # this branch is satisfied outright
    if any_unparseable:
        UNPARSEABLE_LOGIC.add(logic.strip())
        return True, False        # surfaced, but flagged as uncertain
    return False, True


def eval_branching(logic: str, row: dict) -> bool:
    """Whether a field applies. Kept for callers that don't need the certainty flag."""
    return evaluate_branching(logic, row)[0]


def report_unparseable_logic(stream=sys.stderr) -> None:
    """Say once, at the end of a run, which branching logic couldn't be understood."""
    if not UNPARSEABLE_LOGIC:
        return
    print(
        f"\nNote: {len(UNPARSEABLE_LOGIC)} branching condition(s) couldn't be fully understood.\n"
        "Fields controlled by them were included in the worklist and marked as 'check whether\n"
        "this applies' rather than left out — but do check them, and pass this list on so the\n"
        "tool can be taught to read them:",
        file=stream,
    )
    for logic in sorted(UNPARSEABLE_LOGIC):
        print(f"    {logic}", file=stream)


def extract_branching_triggers(logic: str) -> list:
    if not logic:
        return []
    out, seen = [], set()
    for clause in re.split(r"\s+(?:AND|OR)\s+", logic, flags=re.IGNORECASE):
        parts = _clause_parts(clause)
        if parts and parts[0] not in seen:
            seen.add(parts[0]); out.append(parts[0])
    return out


def augment_fields_with_triggers(fields: list, metadata: list) -> tuple[list, dict]:
    """Pull in any gate fields referenced by branching logic on `fields`."""
    meta_by = {m["field_name"]: m for m in metadata}
    prereq, augmented, seen = {}, list(fields), {f.split("___")[0] for f in fields}
    for f in fields:
        base = f.split("___")[0]
        m = meta_by.get(base)
        if not m:
            continue
        triggers = extract_branching_triggers(m.get("branching_logic", "") or "")
        if triggers:
            prereq[base] = triggers
        for t in triggers:
            if t not in seen:
                augmented.append(t); seen.add(t)
    return augmented, prereq


def logic_is_readable(logic: str) -> bool:
    """Whether every clause of this branching condition parses at all.

    Drives the wording of the RA-facing prerequisite row. An unreadable condition (a datediff
    call, say) used to be printed raw after the words "only if", which reads as an instruction
    written in a language the RA doesn't speak — and implies we understood it. We didn't, and
    the cells below are amber for exactly that reason, so the row should say so.
    """
    if not logic or not logic.strip():
        return True
    for clause in re.split(r"\s+(?:AND|OR)\s+", logic, flags=re.IGNORECASE):
        if _clause_parts(clause) is None:
            return False
    return True


def prereq_text(logic: str, meta_by: dict) -> str:
    """The second header row's cell for one field: its prerequisite, in plain English."""
    logic = (logic or "").strip()
    if not logic:
        return ""
    if logic_is_readable(logic):
        return "only if " + format_branching_logic(logic, meta_by)
    return "couldn't read this condition: " + logic


def format_branching_logic(logic: str, meta_by: dict) -> str:
    """Render branching logic with choice labels for the RA-facing prereq row."""
    if not logic:
        return ""
    parts = re.split(r"(\s+(?:AND|OR)\s+)", logic, flags=re.IGNORECASE)
    out = []
    for tok in parts:
        if re.fullmatch(r"\s+(?:AND|OR)\s+", tok, flags=re.IGNORECASE):
            out.append(tok.strip().upper()); continue
        parts = _clause_parts(tok)
        if not parts:
            out.append(tok.strip()); continue
        field, choice_code, op, val = parts
        choices = parse_choices(meta_by.get(field, {}).get("select_choices_or_calculations", "") or "")
        if choice_code is not None:
            label = choices.get(choice_code, choice_code)
            verb = "includes" if val == "1" else "excludes"
            out.append(f"{field} {verb} {label}")
        else:
            label = choices.get(val, val)
            out.append(f"{field} {op} {label}")
    return " ".join(out)


def display_headers(fields: list, label_map: dict, taken=()) -> list:
    """Human column headings for `fields`, guaranteed distinct — display only, never a key.

    A field's LABEL is what the RA reads; the field NAME is what identifies it. REDCap only
    requires names to be unique, and shared labels are completely ordinary: a real 160-field
    colorectal data dictionary had 44 labels used by more than one field ("Date", "Other,
    specify", "Result"). The builder used to rename the dataframe's columns to labels and then
    read each cell back BY LABEL — with a shared label pandas returned a two-row Series instead
    of a value, openpyxl refused it ("Cannot convert ... to Excel"), and the build died outright.
    Every lookup keys on the field name now.

    This function exists for the other half of that problem: two columns under one heading are
    unreadable for the RA even when the machinery copes. The first field to use a label keeps it
    plain; anything after it gets its field name in parentheses, so the heading says which field
    it is. `taken` seeds the used-headings set (the ID columns sit to the left of these).
    """
    used = {str(t) for t in taken}
    out = []
    for f in fields:
        label = str(label_map.get(f, f))
        header = label if label not in used else f"{label} ({f})"
        n = 2
        while header in used:            # pathological: a plain label already reads like that
            header = f"{label} ({f} #{n})"
            n += 1
        used.add(header)
        out.append(header)
    return out


def labelize(df: pd.DataFrame, metadata: list, fields: list) -> tuple[pd.DataFrame, dict]:
    """Return (labeled_df, label_map: raw_field -> human header).

    The dataframe's COLUMNS keep their field names. Only the values are turned into labels
    (choice codes into choice text, ticked checkbox options into a joined list); the label map
    is handed to the writer for the header row and used nowhere else.
    """
    meta_by = {m["field_name"]: m for m in metadata}
    out = df.copy()
    bases, seen = [], set()
    for f in fields:
        b = f.split("___")[0]
        if b not in seen and b in meta_by:
            bases.append(b); seen.add(b)
    label_map = {}
    for base in bases:
        m = meta_by[base]
        ftype = m.get("field_type", "")
        choices = parse_choices(m.get("select_choices_or_calculations", ""))
        label = clean_label(m.get("field_label", "")) or base

        if ftype == "checkbox":
            cols = [c for c in out.columns if c == base or c.startswith(f"{base}___")]
            if not cols:
                continue

            def _code(col, b=base):
                suf = col[len(b):]
                if suf.startswith("____"): return "-" + suf[4:]
                if suf.startswith("___"):  return suf[3:]
                return ""

            def _labels(row, cols=cols, choices=choices, _code=_code):
                return ", ".join(choices.get(_code(c), _code(c))
                                 for c in cols if str(row.get(c, "")) == "1")

            out[base] = out.apply(_labels, axis=1)
            out = out.drop(columns=[c for c in cols if c != base])
        elif ftype in ("radio", "dropdown", "yesno", "truefalse") and base in out.columns and choices:
            out[base] = out[base].map(lambda v: choices.get(str(v), v))

        label_map[base] = label
    return out, label_map


# -----------------------------------------------------------------------------
# Worklist builder
# -----------------------------------------------------------------------------

SENTINEL_CODES = {"666", "-666", "-777", "-888", "-999"}

YELLOW          = PatternFill(start_color=YELLOW_HEX, end_color=YELLOW_HEX, fill_type="solid")
# A distinct fill for cells whose branching condition we could not read. These are shown so
# nothing is ever silently dropped, but they are NOT an assertion that the RA missed something —
# they mean "we couldn't tell whether this applies; please check".
UNCERTAIN       = PatternFill(start_color=AMBER_HEX, end_color=AMBER_HEX, fill_type="solid")
HEADER_FILL     = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
RESPONSE_HEADER = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
HEADER_FONT     = Font(bold=True, color="FFFFFF")
PREREQ_FONT     = Font(italic=True, size=9, color="666666")


def _option_code(col: str, base: str) -> str:
    suf = col[len(base):]
    if suf.startswith("____"): return "-" + suf[4:]
    if suf.startswith("___"):  return suf[3:]
    return ""


def missing_mask_for(field: str, raw_row: pd.Series, meta_by: dict,
                      flag_sentinels: bool = True) -> bool:
    """Whether this cell should be flagged. See missing_and_certainty_for for the certainty flag."""
    return missing_and_certainty_for(field, raw_row, meta_by, flag_sentinels)[0]


def missing_and_certainty_for(field: str, raw_row: pd.Series, meta_by: dict,
                              flag_sentinels: bool = True) -> "tuple[bool, bool]":
    """(should_flag, certain). certain=False means we could not read the field's condition."""
    m = meta_by.get(field)
    if not m:
        return False, True
    applies, certain = evaluate_branching(m.get("branching_logic") or "", raw_row.to_dict())
    if not applies:
        return False, certain
    ftype = m.get("field_type", "")
    if ftype == "checkbox":
        ticked = [
            _option_code(k, field)
            for k in raw_row.index
            if (k == field or k.startswith(f"{field}___")) and str(raw_row.get(k, "")) == "1"
        ]
        if not ticked:
            return True, certain
        return flag_sentinels and all(t in SENTINEL_CODES for t in ticked), certain
    val = str(raw_row.get(field, "")).strip()
    if val == "":
        return True, certain
    return flag_sentinels and val in SENTINEL_CODES, certain


def build_workbook(labeled: pd.DataFrame, raw_by_id: dict, meta_by: dict,
                    prereq_map: dict, fields: list, label_map: dict,
                    id_cols: list, title: str, flag_sentinels: bool) -> Workbook | None:
    """One workbook = wide table, applicable-but-blank cells in yellow."""
    rows_with_work, fields_with_work = [], set()
    join_id = id_cols[0]
    for _, lrow in labeled.iterrows():
        rid = str(lrow.get(join_id, "")).strip()
        rrow = raw_by_id.get(rid)
        if rrow is None:
            continue
        missing, uncertain = [], set()
        for f in fields:
            flag, certain = missing_and_certainty_for(f, rrow, meta_by, flag_sentinels)
            if flag:
                missing.append(f)
                if not certain:
                    uncertain.add(f)
        if missing:
            rows_with_work.append((lrow, rrow, set(missing), uncertain))
            fields_with_work.update(missing)

    if not rows_with_work:
        return None

    # Gate context — surface fields that *gate* something flagged. Gate columns come first so
    # the RA reads "why is this flagged?" left to right, and they are ordered by the DATA
    # DICTIONARY's own field order: `meta_by` is built from the metadata export in file order,
    # so its key order IS DD order.
    #
    # This used to iterate `context_set` directly and insert each gate at the front. Set order
    # over strings depends on PYTHONHASHSEED, so a workbook with two or more gate columns laid
    # them out differently on different runs of the same command against the same data:
    # round-to-round worklists weren't diffable, RAs saw columns move, and nothing built
    # through here could be byte-reproducible.
    context_set = {g for f in fields_with_work for g in prereq_map.get(f, [])
                   if g not in fields_with_work}
    dd_order = {name: i for i, name in enumerate(meta_by)}
    gates = sorted(context_set, key=lambda g: (dd_order.get(g, len(dd_order)), g))
    display_set = fields_with_work | context_set
    display_fields = []
    for f in gates + [f for f in fields if f in display_set]:
        if f not in display_fields:
            display_fields.append(f)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header = id_cols + display_headers(display_fields, label_map, taken=id_cols) + ["RESPONSE"]
    prereq_row = (
        [""] * len(id_cols)
        + [prereq_text(meta_by.get(f, {}).get("branching_logic") or "", meta_by)
           for f in display_fields]
        + ["per-row notes from the RA: why blank, RESOLVED, etc."]
    )
    ws.append(header); ws.append(prereq_row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    # RESPONSE column header — distinct color so RAs notice it
    ws.cell(row=1, column=len(header)).fill = RESPONSE_HEADER
    for cell in ws[2]:
        cell.font = PREREQ_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = f"{get_column_letter(len(id_cols) + 1)}3"

    for lrow, rrow, missing, uncertain in rows_with_work:
        out = [lrow.get(c, "") for c in id_cols]
        for f in display_fields:
            # By field NAME. Reading the cell back by its label is what crashed the builder on
            # every data dictionary that reuses a label — see display_headers().
            out.append(lrow.get(f, ""))
        out.append("")  # empty RESPONSE cell for the RA
        ws.append(out)
        row_idx = ws.max_row
        for i, f in enumerate(display_fields):
            if f in missing:
                ws.cell(row=row_idx, column=len(id_cols) + 1 + i).fill = (
                    UNCERTAIN if f in uncertain else YELLOW)

    for i in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(32, len(str(header[i-1])) + 4))
    # RESPONSE column wider for free-text
    ws.column_dimensions[get_column_letter(len(header))].width = 50
    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 24
    return wb


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

def _expand_checkbox_columns(fields: list, raw_cols: list, meta_by: dict) -> list:
    out, seen = [], set()
    for f in fields:
        base = f.split("___")[0]
        m = meta_by.get(base)
        if m and m.get("field_type") == "checkbox":
            for col in raw_cols:
                if (col == base or col.startswith(f"{base}___")) and col not in seen:
                    out.append(col); seen.add(col)
        elif f in raw_cols and f not in seen:
            out.append(f); seen.add(f)
    return out


# REDCap "Download Data Dictionary" CSV (human headers) -> API metadata keys, for no-token mode.
_DD_TO_META = {
    "Variable / Field Name": "field_name", "Form Name": "form_name", "Section Header": "section_header",
    "Field Type": "field_type", "Field Label": "field_label",
    "Choices, Calculations, OR Slider Labels": "select_choices_or_calculations", "Field Note": "field_note",
    "Text Validation Type OR Show Slider Number": "text_validation_type_or_show_slider_number",
    "Text Validation Min": "text_validation_min", "Text Validation Max": "text_validation_max",
    "Identifier?": "identifier", "Branching Logic (Show field only if...)": "branching_logic",
    "Required Field?": "required_field", "Custom Alignment": "custom_alignment",
    "Question Number (surveys only)": "question_number", "Matrix Group Name": "matrix_group_name",
    "Matrix Ranking?": "matrix_ranking", "Field Annotation": "field_annotation",
}


def load_metadata_csv(path):
    """Load a local metadata CSV (no-token mode). Accepts either a REDCap API metadata export
    (already has 'field_name') or a Designer 'Download Data Dictionary' CSV (human headers)."""
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    if "field_name" not in df.columns:
        df = df.rename(columns=_DD_TO_META)
    return df.to_dict("records")


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--url", help="REDCap API URL. Only needed if REDCAP_URL isn't already "
                                 "in your ARGO settings file")
    ap.add_argument("--token-env", help="Name of the setting holding your access key for this "
                                        "study, e.g. CRC_TOKEN (access-key mode)")
    ap.add_argument("--records-csv", help="No-token mode: local record export CSV (use instead of --url/--token-env)")
    ap.add_argument("--metadata-csv", help="No-token mode: local Data Dictionary CSV (paired with --records-csv)")
    ap.add_argument("--fields", required=True, help="Path to fields YAML")
    ap.add_argument("--out", required=True, help="Output directory (a per-round subdir will be appended)")
    ap.add_argument("--round", dest="round_tag", default="",
                    help="Round label appended to --out so reruns don't overwrite. Defaults to today's date (YYYY-MM-DD). Pass empty string '--round=' to write to --out directly (legacy behavior).")
    ap.add_argument("--scope-ids", help="Optional CSV of record IDs to restrict to (one column, first row is header)")
    ap.add_argument("--id-field", default="record_id", help="REDCap record-id field name")
    ap.add_argument("--extra-id-cols", default="", help="Comma-separated extra ID columns to include in the worklist")
    args = ap.parse_args()
    if args.round_tag is None:
        args.round_tag = ""
    # Default round tag = today's date; explicit '' disables the subdir.
    if args.round_tag == "" and "--round=" not in " ".join(sys.argv) and "--round" not in sys.argv:
        args.round_tag = _dt.date.today().isoformat()
    if args.round_tag:
        args.out = os.path.join(args.out, args.round_tag)
        print(f"Round: {args.round_tag} → writing to {args.out}/")

    cfg = yaml.safe_load(open(args.fields))
    workbooks_cfg = cfg.get("workbooks") or []
    if not workbooks_cfg:
        sys.exit(f"No `workbooks:` in {args.fields}")

    os.makedirs(args.out, exist_ok=True)
    extra_id_cols = [c.strip() for c in args.extra_id_cols.split(",") if c.strip()]
    id_cols = [args.id_field] + extra_id_cols

    if args.records_csv:
        if not args.metadata_csv:
            sys.exit(
                "To build worklists from files on your computer, I need two of them: the records\n"
                "export and the data dictionary. You've given me only the records.\n"
                "\n"
                "Download the data dictionary from the REDCap project's Data Dictionary page and\n"
                "pass it as well:\n"
                "\n"
                "    --records-csv records.csv --metadata-csv datadictionary.csv"
            )
        print(f"No-token mode: reading {args.records_csv} + {args.metadata_csv} ...")
        raw = pd.read_csv(args.records_csv, dtype=str, keep_default_na=False)
        metadata = load_metadata_csv(args.metadata_csv)
    elif args.token_env:
        # Load the ARGO settings file ourselves. There is nothing for the user to source first,
        # and --url is only needed if their REDCap address isn't in that file already.
        load_env_file()
        url = args.url or os.environ.get("REDCAP_URL")
        token = os.environ.get(args.token_env)
        if not token:
            sys.exit(
        f"No access key called {args.token_env} is in your ARGO settings file, so I can't reach\n"
        "REDCap.\n"
        "\n"
        "An access key (REDCap calls it an API token) is a long password that lets a tool read\n"
        "or update one specific REDCap project on your behalf. Your REDCap administrator\n"
        "creates it for you — it isn't something you can generate yourself.\n"
        "\n"
        "If you already have one: open your ARGO folder, double-click 'Add keys here' to open\n"
        f"your settings file, add the line {args.token_env}=<your key>, save, and run this again.\n"
        "Never type a key into a chat message — it stays in the transcript.\n"
        "\n"
        "No key at all? Nothing is blocked: download the records export and the data dictionary\n"
        "from the REDCap website and pass them with --records-csv and --metadata-csv instead."
    )
        if not url:
            sys.exit(
        "I have your access key, but not the web address of your REDCap system.\n"
        "\n"
        "It's a single line of text ending in /api/. Open your ARGO folder, double-click\n"
        "'Add keys here', and add it on the REDCAP_URL line:\n"
        "\n"
        "    REDCAP_URL=https://your-redcap-site.org/api/"
    )
        print(f"Pulling records + metadata from {url} ...")
        raw = pull_records(url, token)
        metadata = pull_metadata(url, token)
    else:
        sys.exit(
            "I need to know where to get the study's data from, and you haven't told me yet.\n"
            "There are two ways, and you only need one:\n"
            "\n"
            "  Working from files you downloaded from the REDCap website (the usual way):\n"
            "      --records-csv records.csv --metadata-csv datadictionary.csv\n"
            "\n"
            "  Or, if you have an access key for this study set up:\n"
            "      --token-env YOUR_STUDY_TOKEN"
        )
    meta_by = {m["field_name"]: m for m in metadata}

    if args.id_field not in raw.columns:
        sys.exit(f"--id-field {args.id_field!r} not in record columns: {list(raw.columns)[:8]}...")

    if args.scope_ids:
        scope = pd.read_csv(args.scope_ids, dtype=str)
        ids = set(scope.iloc[:, 0].astype(str).str.strip())
        raw = raw[raw[args.id_field].astype(str).str.strip().isin(ids)]
        print(f"  scoped to {len(raw)} record(s) via {args.scope_ids}")

    if "redcap_data_access_group" not in raw.columns:
        raw["redcap_data_access_group"] = "unassigned"
    raw["redcap_data_access_group"] = raw["redcap_data_access_group"].replace("", "unassigned").fillna("unassigned")

    raw_by_id = {str(r[args.id_field]).strip(): r for _, r in raw.iterrows()}
    dags = sorted(raw["redcap_data_access_group"].unique())
    print(f"DAGs: {dags}")

    for wb_cfg in workbooks_cfg:
        wb_name = wb_cfg["name"]
        title = wb_cfg.get("title", wb_name)
        fields = wb_cfg["fields"]

        augmented, prereq_map = augment_fields_with_triggers(fields, metadata)
        expanded = _expand_checkbox_columns(augmented, list(raw.columns), meta_by)
        keep_cols = [c for c in id_cols if c in raw.columns] + ["redcap_data_access_group"] + \
                    [c for c in expanded if c not in id_cols]
        sub_raw = raw[keep_cols].copy()
        # Values become human-readable; COLUMN NAMES stay field names. They used to be renamed
        # to labels here, which quietly made the label the row's key — and two fields sharing a
        # label then produced two columns with the same name and a crash downstream.
        labeled, label_map = labelize(sub_raw, metadata, augmented)

        for dag in dags:
            sub = labeled[labeled["redcap_data_access_group"] == dag]
            if sub.empty:
                continue
            for flag_sentinels, subfolder in [(True, "with_MDC"), (False, "no_MDC")]:
                wb = build_workbook(sub, raw_by_id, meta_by, prereq_map,
                                    fields, label_map, id_cols, title, flag_sentinels)
                folder = os.path.join(args.out, subfolder)
                os.makedirs(folder, exist_ok=True)
                tag = f"{wb_name}_{dag}"
                if wb is None:
                    print(f"  skip  {subfolder}/{tag}: nothing to review")
                    continue
                path = os.path.join(folder, f"{tag}.xlsx")
                wb.save(path)
                ws = wb.active
                print(f"  wrote {path}  ({ws.max_row - 2} patients × {ws.max_column - len(id_cols)} fields)")

    print("Done.")
    report_unparseable_logic()


if __name__ == "__main__":
    main()
