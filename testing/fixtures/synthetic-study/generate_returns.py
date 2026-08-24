#!/usr/bin/env python3
"""Seeded generator for RA-RETURNED worklists on the SYN synthetic study.

SYNTHETIC TEST STUDY — every value written here is fabricated. No real people,
sites, or data. This is the input fixture for the QA specialist's **task 2**
(audit what the RAs send back), which otherwise had no test coverage at all.

What it does, deterministically, from the committed fixture:

  1. Runs `build_worklists.py` on records.csv + datadictionary.csv + qa_fields.yaml
     into `<out>/build/` (with `--round=` so the path has no date in it).
     With `--from-worklists DIR` it SKIPS that build and reads the worklists already
     in DIR instead — the point being that a live session's own build (its own
     workbooks, fields and sites, from the DD-driven config the skill writes) can be
     turned into matching RA returns. Without a hand-tuned plan for those workbooks
     the edit budget is derived from each one (see derive_plan). The default path is
     unchanged and stays byte-for-byte reproducible.
  2. Takes the four `with_MDC/` workbooks (2 workbooks x 2 DAGs) and writes an
     RA-RETURNED copy of each into `<out>/returned/`, with ENGINEERED edits:

       filled_value   yellow cells filled with a plausible in-choice-list value
       filled_mdc     yellow cells filled with an MDC sentinel code (-666/-777/-888/-999)
       untouched      yellow cells left exactly as sent
       notes_on_changed  rows that got a fill AND an RA note in the RESPONSE column
       notes_only        rows with an RA note but NO cell change  (the "VERIFY" case)
       out_of_scope   cells the RA edited that were NEVER flagged (gate-context columns)
       amber_filled   amber "we couldn't read this condition" cells the RA filled

  3. Writes `<out>/returned/returned_counts.json` with the exact counts, and with
     `--update-manifest` merges the same numbers into MANIFEST.json's `returned`
     block (leaving every other key untouched).

No .xlsx is ever committed: `tests/test_qa_audit_round_trip.py` runs this into a
temp directory and asserts what `review_responses.py` reports against MANIFEST.

NOTE: `generate.py` rewrites MANIFEST.json from scratch and does not know about
the `returned` block. If you rerun `generate.py`, rerun this with
`--update-manifest` afterwards. The round-trip test fails loudly if the block is
missing or stale — it is never silently wrong.

Run:
    python3 generate_returns.py --out /tmp/syn-returns [--update-manifest]
    python3 generate_returns.py --out /tmp/live-returns \\
        --from-worklists qa-specialist/<study>/worklists/<round>

Requires pandas / openpyxl / pyyaml (same dependencies build_worklists.py needs).
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import random
import re
import subprocess
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, "..", "..", ".."))
QA_SKILL = os.path.join(REPO, "plugins", "argo-qa-specialist", "skills", "qa-worklists")
BUILDER = os.path.join(QA_SKILL, "build_worklists.py")

SEED = 20260820

# The fills the builder paints, imported from the skill rather than retyped. A fixture that
# hunts for a colour the builder stopped using would engineer zero RA answers and assert
# nothing, cheerfully and green. See qa_colours.py.
sys.path.insert(0, QA_SKILL)
from qa_colours import AMBER_HEX, YELLOW_HEX  # noqa: E402
SENTINELS = ["-888", "-777", "-999", "-666"]
RESPONSE_HEADER_TOKENS = ("response", "comment", "note")

# The with_MDC workbooks this fixture engineers, and the exact edit budget for each.
# Sized so every count fits inside the smallest workbook (8 yellow cells, 8 rows).
PLAN = {
    "clinical_core_site_alpha": {
        "filled_value": 8, "filled_mdc": 5, "notes_on_changed": 4,
        "notes_only": 3, "out_of_scope": 3, "amber_filled": 3,
    },
    "clinical_core_site_beta": {
        "filled_value": 5, "filled_mdc": 3, "notes_on_changed": 3,
        "notes_only": 2, "out_of_scope": 2, "amber_filled": 2,
    },
    "demo_followup_site_alpha": {
        "filled_value": 4, "filled_mdc": 2, "notes_on_changed": 2,
        "notes_only": 2, "out_of_scope": 2, "amber_filled": 0,
    },
    "demo_followup_site_beta": {
        "filled_value": 2, "filled_mdc": 2, "notes_on_changed": 2,
        "notes_only": 1, "out_of_scope": 2, "amber_filled": 0,
    },
}

# Synthetic RA wording. Deliberately mixed: clean "RESOLVED" markers, an
# explanation that warrants no recode, and one vague note.
NOTES_ON_CHANGED = [
    "Checked the synthetic chart and entered these in REDCap.",
    "Values found in the fake source folder; REDCap updated.",
    "Entered in REDCap — please re-pull to confirm.",
    "Filled from the test source sheet, nothing outstanding now.",
]
NOTES_ONLY = [
    "RESOLVED",
    "Synthetic participant transferred; no chart held at this site.",
    "RESOLVED in REDCap already, spreadsheet not updated.",
]


# ---------------------------------------------------------------------------
# Data dictionary helpers (stdlib — mirrors build_worklists' label conventions)
# ---------------------------------------------------------------------------

def clean_label(label: str) -> str:
    """Same normalisation build_worklists.clean_label applies to header labels."""
    if not label:
        return label
    return re.sub(r"\s+", " ", label).strip().rstrip(" ?.:;")


def parse_choices(raw: str) -> dict:
    out = {}
    for part in (raw or "").split("|"):
        if "," in part:
            code, label = part.split(",", 1)
            out[code.strip()] = label.strip()
    return out


def load_dd(path: str) -> dict:
    """cleaned field_label -> {'field': name, 'choices': {code: label}}"""
    by_label = {}
    with open(path, newline="") as fh:
        for row in csv.DictReader(fh):
            label = clean_label(row.get("field_label", ""))
            if not label:
                continue
            by_label.setdefault(label, {
                "field": row["field_name"],
                "choices": parse_choices(row.get("select_choices_or_calculations", "")),
            })
    return by_label


def plausible_value(header: str, current, dd_by_label: dict, rng) -> str:
    """A value an RA could plausibly have typed: a real choice label for the
    field, never an MDC label, and never equal to what is already in the cell."""
    spec = dd_by_label.get(str(header).strip(), {})
    choices = spec.get("choices") or {}
    options = [lbl for code, lbl in choices.items() if code not in SENTINELS]
    cur = "" if current is None else str(current).strip()
    options = [o for o in options if o != cur]
    if options:
        return options[rng.randrange(len(options))]
    # text/date fields: the SKILL's documented "filled" marker. If that is somehow already
    # what the cell says, vary it — an "edit" that changes nothing is not an edit, and every
    # count below assumes the value moved.
    return "filled" if cur != "filled" else "filled (rechecked)"


# ---------------------------------------------------------------------------
# Workbook edits
# ---------------------------------------------------------------------------

def _fill_hex(cell) -> str:
    f = cell.fill
    if not f or not f.fgColor:
        return ""
    return str(f.fgColor.rgb or "").upper()


def _response_col(headers) -> int | None:
    for i, h in enumerate(headers, 1):
        if h and any(t in str(h).lower() for t in RESPONSE_HEADER_TOKENS):
            return i
    return None


def scan_workbook(ws, headers, resp_col) -> tuple:
    """Classify every data cell of a built worklist.

    Returns (yellow, amber, plain, data_rows). Cells are keyed by (row, HEADER, column, id)
    rather than by column index on purpose: which column a gate-context field lands in is a
    property of the build, and this generator's picks must not depend on it.
    """
    yellow, amber, plain, data_rows = [], [], [], []
    for r in range(3, ws.max_row + 1):
        rid = str(ws.cell(row=r, column=1).value or "").strip()
        if not rid:
            continue
        data_rows.append(r)
        for c in range(2, ws.max_column + 1):
            if c == resp_col:
                continue
            cell = ws.cell(row=r, column=c)
            hexv = _fill_hex(cell)
            entry = (r, str(headers[c - 1]), c, rid)
            if hexv.endswith(YELLOW_HEX):
                yellow.append(entry)
            elif hexv.endswith(AMBER_HEX):
                amber.append(entry)
            elif cell.value not in (None, ""):
                plain.append(entry)
    return yellow, amber, plain, data_rows


def derive_plan(yellow: list, amber: list, plain: list, data_rows: list) -> dict:
    """An edit budget for a worklist this generator has never seen before.

    `--from-worklists` aims the kit at whatever a live session actually built, where the
    hand-tuned PLAN above does not apply — different field sets, different sites, different
    row counts. Take a fixed share of what is actually there, and always leave at least one
    yellow cell untouched: "the RA answered everything" is the one return shape that
    exercises nothing. Every request here is clamped to what the workbook can supply.
    """
    n = len(yellow)
    filled_value = min(max(n // 2, 1) if n else 0, 8)
    filled_mdc = min(n // 4, 5)
    while n and filled_value + filled_mdc >= n:
        if filled_mdc:
            filled_mdc -= 1
        elif filled_value > 1:
            filled_value -= 1
        else:
            break
    return {
        "filled_value": filled_value,
        "filled_mdc": filled_mdc,
        "notes_on_changed": 4,          # clamped to the rows that actually changed
        "notes_only": 3,                # clamped to the rows that did not
        "out_of_scope": 3,              # clamped to the unflagged filled cells available
        "amber_filled": min(len(amber), 3),
    }


def _take(requested: int, available: int, what: str, src_path: str, strict: bool) -> int:
    """How many of `requested` we can actually do.

    The committed fixture runs strict: its PLAN is hand-sized to the fixture and a shortfall
    means the fixture or the builder changed, which must fail loudly rather than quietly
    produce a different test input. A derived plan (`--from-worklists`) clamps instead.
    """
    if requested <= available:
        return requested
    if strict:
        raise SystemExit(f"{src_path}: plan wants {requested} {what}, "
                         f"only {available} available")
    return available


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def engineer_return(src_path: str, dst_path: str, plan: "dict | None", dd_by_label: dict,
                    seed: int, strict: bool = True) -> dict:
    """Write an RA-returned copy of `src_path` and return its exact edit counts.

    `plan=None` derives the edit budget from the workbook itself (see derive_plan).
    """
    from openpyxl import load_workbook

    wb = load_workbook(src_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    resp_col = _response_col(headers)
    if resp_col is None:
        raise SystemExit(f"{src_path}: no RESPONSE column — build_worklists should add one")

    yellow, amber, plain, data_rows = scan_workbook(ws, headers, resp_col)
    if plan is None:
        plan = derive_plan(yellow, amber, plain, data_rows)
        strict = False

    rng = random.Random(seed)

    n_fill = plan["filled_value"] + plan["filled_mdc"]
    if n_fill > len(yellow):
        if strict:
            raise SystemExit(f"{src_path}: plan wants {n_fill} fills, "
                             f"only {len(yellow)} yellow cells")
        n_fill = len(yellow)

    order = sorted(yellow)
    rng.shuffle(order)
    n_value = min(plan["filled_value"], n_fill)
    to_value = sorted(order[:n_value])
    to_mdc = sorted(order[n_value:n_fill])

    # Fills are counted by what actually MOVED: a "fill" that writes back the value already in
    # the cell is not an edit and no audit could ever report it.
    changed_ids = set()
    n_value_changed = 0
    for r, header, c, rid in to_value:
        cell = ws.cell(row=r, column=c)
        before = _text(cell.value)
        cell.value = plausible_value(header, cell.value, dd_by_label, rng)
        if _text(cell.value) != before:
            n_value_changed += 1
            changed_ids.add(rid)
    mdc_on_mdc_field = 0
    n_mdc_changed = 0
    for i, (r, header, c, rid) in enumerate(to_mdc):
        code = SENTINELS[i % len(SENTINELS)]
        spec = dd_by_label.get(header.strip(), {})
        if code in (spec.get("choices") or {}):
            mdc_on_mdc_field += 1
        cell = ws.cell(row=r, column=c)
        before = _text(cell.value)
        cell.value = code
        if _text(cell.value) != before:
            n_mdc_changed += 1
            changed_ids.add(rid)

    changed_rows = sorted({e[0] for e in to_value + to_mdc})

    # Notes on rows that DID change something.
    n_notes = _take(plan["notes_on_changed"], len(changed_rows),
                    "notes on changed rows", src_path, strict)
    note_rows = changed_rows[:n_notes]
    for i, r in enumerate(note_rows):
        ws.cell(row=r, column=resp_col).value = NOTES_ON_CHANGED[i % len(NOTES_ON_CHANGED)]

    # Notes on rows with NO cell change at all — the "RA said RESOLVED but the
    # cell is still blank" case the audit is supposed to surface as VERIFY.
    untouched_rows = [r for r in data_rows if r not in set(changed_rows)]
    n_note_only = _take(plan["notes_only"], len(untouched_rows),
                        "note-only rows", src_path, strict)
    note_only_rows = untouched_rows[:n_note_only]
    for i, r in enumerate(note_only_rows):
        ws.cell(row=r, column=resp_col).value = NOTES_ONLY[i % len(NOTES_ONLY)]

    # Out-of-scope edits: cells that were never flagged, on rows that are
    # otherwise silent. A correct audit reports these separately — they are not
    # answers to anything we asked.
    used_rows = set(changed_rows) | set(note_only_rows)
    candidates = sorted(p for p in plain if p[0] not in used_rows)
    n_oos = _take(plan["out_of_scope"], len(candidates), "out-of-scope edits", src_path, strict)
    rng.shuffle(candidates)
    oos = sorted(candidates[:n_oos])
    oos_detail = []
    for r, header, c, rid in oos:
        cell = ws.cell(row=r, column=c)
        before = _text(cell.value)
        cell.value = plausible_value(header, cell.value, dd_by_label, rng)
        if _text(cell.value) == before:
            continue
        oos_detail.append({"record": rid, "column": header,
                           "was": before, "now": str(cell.value)})

    # Amber ("please check whether this applies") cells the RA filled in. An answer in an
    # amber cell is still an answer, and the audit must report it as one.
    n_amber = _take(plan["amber_filled"], len(amber), "amber cells to fill", src_path, strict)
    amber_pick = sorted(amber)
    rng.shuffle(amber_pick)
    amber_pick = sorted(amber_pick[:n_amber])
    amber_ids = []
    for r, header, c, rid in amber_pick:
        cell = ws.cell(row=r, column=c)
        before = _text(cell.value)
        cell.value = plausible_value(header, cell.value, dd_by_label, rng)
        if _text(cell.value) == before:
            continue
        amber_ids.append(rid)

    # What review_responses.py should now report. Answers = yellow + amber; the note-only
    # bucket is whatever kept a note and never got an answer.
    answered_ids = sorted(changed_ids | set(amber_ids))
    note_only_ids = sorted(
        rid for rid in (str(ws.cell(row=r, column=1).value).strip() for r in note_only_rows)
        if rid not in set(answered_ids))

    os.makedirs(os.path.dirname(dst_path) or ".", exist_ok=True)
    wb.save(dst_path)

    return {
        "source_worklist": os.path.basename(src_path),
        "returned_file": os.path.basename(dst_path),
        "yellow_cells_total": len(yellow),
        "amber_cells_total": len(amber),
        "filled_with_value": n_value_changed,
        "filled_with_mdc": n_mdc_changed,
        "filled_with_mdc_on_mdc_enabled_field": mdc_on_mdc_field,
        "left_untouched": len(yellow) - n_fill,
        "notes_on_changed_rows": len(note_rows),
        "notes_only_rows": len(note_only_rows),
        "out_of_scope_edits": len(oos_detail),
        "amber_cells_filled": len(amber_ids),
        "records_with_changed_cells": len(changed_ids),
        "records_with_answers": len(answered_ids),
        "changed_cells_total": n_value_changed + n_mdc_changed,
        "note_only_record_ids": note_only_ids,
        "out_of_scope_detail": oos_detail,
        "amber_filled_record_ids": sorted(amber_ids),
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run_builder(out_root: str) -> str:
    build_dir = os.path.join(out_root, "build")
    proc = subprocess.run(
        [sys.executable, BUILDER,
         "--records-csv", os.path.join(HERE, "records.csv"),
         "--metadata-csv", os.path.join(HERE, "datadictionary.csv"),
         "--fields", os.path.join(HERE, "qa_fields.yaml"),
         "--out", build_dir, "--id-field", "syn_id", "--round="],
        capture_output=True, text=True, timeout=600,
    )
    if proc.returncode != 0:
        raise SystemExit("build_worklists.py failed:\n" + proc.stdout + proc.stderr)
    return build_dir


def find_worklists(root: str, variant: str = "with_MDC") -> list:
    """The .xlsx worklists inside a directory build_worklists.py wrote.

    Accepts any of the shapes a real run leaves behind:
        <root>/with_MDC/*.xlsx          the builder's output directory
        <root>/<round>/with_MDC/*.xlsx  the builder's default, per-round subdir
        <root>/*.xlsx                   a folder of workbooks copied out by hand
    `variant` picks with_MDC (flags MDC sentinels too) or no_MDC.
    """
    root = os.path.abspath(os.path.expanduser(root))
    if not os.path.isdir(root):
        raise SystemExit(
            f"I couldn't find the worklists folder:\n    {root}\n\n"
            "Point --from-worklists at the folder build_worklists.py wrote (the one that\n"
            "contains with_MDC/ and no_MDC/).")
    candidates = []
    direct = os.path.join(root, variant)
    if os.path.isdir(direct):
        candidates.append(direct)
    else:
        for entry in sorted(os.listdir(root)):
            nested = os.path.join(root, entry, variant)
            if os.path.isdir(nested):
                candidates.append(nested)
    if not candidates:
        candidates = [root]
    books = []
    for folder in candidates:
        books += [os.path.join(folder, fn) for fn in sorted(os.listdir(folder))
                  if fn.endswith(".xlsx") and not fn.startswith("~$")]
    if not books:
        raise SystemExit(
            f"No .xlsx worklists under {root} (looked in {variant}/ and the folder itself).\n"
            "Build some first, or point --from-worklists somewhere else.")
    return books


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--out", required=True,
                    help="Directory for the generated build/ and returned/ workbooks")
    ap.add_argument("--from-worklists", dest="from_worklists", default=None,
                    help="Generate returns from worklists that ALREADY exist in this directory "
                         "(a real session's build) instead of building the fixture's own. The "
                         "edit budget is derived from each workbook.")
    ap.add_argument("--variant", default="with_MDC", choices=["with_MDC", "no_MDC"],
                    help="Which subfolder of --from-worklists to read (default with_MDC)")
    ap.add_argument("--seed", type=int, default=SEED)
    ap.add_argument("--update-manifest", action="store_true",
                    help="Merge the counts into MANIFEST.json's `returned` block "
                         "(fixture build only)")
    args = ap.parse_args()

    if args.update_manifest and args.from_worklists:
        raise SystemExit(
            "--update-manifest describes the committed fixture's own build, so it can't be\n"
            "combined with --from-worklists. Run it without --from-worklists to refresh\n"
            "MANIFEST.json.")

    dd_by_label = load_dd(os.path.join(HERE, "datadictionary.csv"))
    returned_dir = os.path.join(args.out, "returned")

    if args.from_worklists:
        sources = find_worklists(args.from_worklists, args.variant)
        print(f"Reading {len(sources)} worklist(s) from {args.from_worklists} ({args.variant}/)")
    else:
        build_dir = run_builder(args.out)
        sources = []
        for name in sorted(PLAN):
            src = os.path.join(build_dir, "with_MDC", f"{name}.xlsx")
            if not os.path.exists(src):
                raise SystemExit(f"expected worklist not produced: {src}")
            sources.append(src)

    per_file, totals = {}, {}
    for i, src in enumerate(sources):
        name = os.path.splitext(os.path.basename(src))[0]
        dst = os.path.join(returned_dir, f"{name}_RETURNED.xlsx")
        counts = engineer_return(src, dst, PLAN.get(name), dd_by_label, args.seed + i,
                                 strict=name in PLAN and not args.from_worklists)
        per_file[name] = counts
        for k, v in counts.items():
            if isinstance(v, int):
                totals[k] = totals.get(k, 0) + v
        print(f"  wrote {dst}  "
              f"({counts['filled_with_value']} values, {counts['filled_with_mdc']} MDC, "
              f"{counts['left_untouched']} untouched, {counts['notes_only_rows']} note-only, "
              f"{counts['out_of_scope_edits']} out-of-scope, "
              f"{counts['amber_cells_filled']} amber)")

    block = {
        "synthetic": True,
        "seed": args.seed,
        "source_mode": args.variant if args.from_worklists else "with_MDC",
        "generator": "generate_returns.py",
        "note": "RA-returned workbooks are generated at test time; no .xlsx is committed.",
        "per_file": per_file,
        "totals": {k: totals[k] for k in sorted(totals)},
        # What review_responses.py should report, derived from the edits above.
        "expected_review_responses": {
            name: {
                "records_with_proposed_updates": c["records_with_answers"],
                "changed_cells": c["changed_cells_total"],
                "note_only_records": len(c["note_only_record_ids"]),
                "response_column_present": True,
                "out_of_scope_edits_detected": c["out_of_scope_edits"],
                "amber_cells_detected": c["amber_cells_filled"],
            }
            for name, c in per_file.items()
        },
    }
    if args.from_worklists:
        block["source_dir"] = os.path.abspath(os.path.expanduser(args.from_worklists))

    os.makedirs(returned_dir, exist_ok=True)
    with open(os.path.join(returned_dir, "returned_counts.json"), "w") as fh:
        json.dump(block, fh, indent=2, sort_keys=True)
        fh.write("\n")

    if args.update_manifest:
        mpath = os.path.join(HERE, "MANIFEST.json")
        manifest = json.load(open(mpath))
        manifest["returned"] = block
        with open(mpath, "w") as fh:
            json.dump(manifest, fh, indent=2, sort_keys=True)
            fh.write("\n")
        print(f"Updated {mpath} → `returned` block")

    print(f"Done. {len(per_file)} returned workbook(s) in {returned_dir}")


if __name__ == "__main__":
    main()
